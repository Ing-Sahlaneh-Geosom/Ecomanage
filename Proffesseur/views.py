from django.contrib.auth.mixins import LoginRequiredMixin
from django.forms import CheckboxSelectMultiple
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, DeleteView , UpdateView , CreateView
from Ecole_admin.utils.mixins import EcoleAssignMixin
from Ecole_admin.form import ProfesseurForm
from Ecole_admin.models import User, Proffeseur , Message
from Ecole_admin.utils.utils import  build_username, unique_username


class ProffeseurList( LoginRequiredMixin , ListView):
    model = Proffeseur
    template_name = 'LesProffesseur.html'
    context_object_name = 'profs'
    login_url = 'Connection'
    paginate_by = 20

    def get_queryset(self):
        queryset = Proffeseur.objects.filter(
            ecole=self.request.user.ecole,
        )

        nom = self.request.GET.get("nom_conplet")
        status = self.request.GET.get("status")
        active = self.request.GET.get("actif")

        if nom and nom != "":
            queryset = queryset.filter(nom_conplet__icontains = nom)
        if status and status != "":
            queryset = queryset.filter(status__icontains=status)
        if active and active != "":
            queryset = queryset.filter(actif__icontains=active)

        return queryset


class ProffeseurDeataille(DetailView):
    model = Proffeseur
    template_name = "detailleProf.html"
    context_object_name = 'Prof'

# views.py
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.urls import reverse
from django.utils.translation import gettext as _
from django.views.generic import CreateView, UpdateView

from Ecole_admin.models import Proffeseur
from Ecole_admin.form import ProfesseurForm
from Ecole_admin.utils.mixins import EcoleAssignMixin

# build_username() et unique_username() doivent déjà exister chez toi
# from .utils import build_username, unique_username
# from accounts.models import User  (ou ton import User)

class AjouterProfesseur(EcoleAssignMixin, CreateView):
    model = Proffeseur
    template_name = "cree_professeur.html"
    form_class = ProfesseurForm

    def form_valid(self, form):
        response = super().form_valid(form)  # self.object = prof créé

        prof = self.object
        created_prof_user = None
        tel = (prof.telephone or "").strip()

        # ✅ Créer user si pas lié
        if not prof.user_id and tel:
            base = build_username(prof.nom_conplet, "prof")
            username = unique_username(User, base)
            email = (prof.email or "").strip() or f"{username}@prof.local"

            with transaction.atomic():
                created_prof_user = User.objects.create(
                    username=username,
                    email=email,
                    nom_complet=prof.nom_conplet,
                    num_tel=tel,
                    role="proffesseur",
                    ecole=prof.ecole,
                    password=make_password(tel)  # mdp = tel
                )
                prof.user = created_prof_user
                prof.save(update_fields=["user"])

        msg = _("Professeur ajouté avec succès ✅")
        if created_prof_user:
            msg += f" | Prof login: {created_prof_user.username} | MDP: {tel}"
        messages.success(self.request, msg)

        return response

    def get_success_url(self):
        return reverse("Lesensiegnants")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["Title"] = _("Ajouter un Professeur")
        context["Submit_text"] = _("Ajouter")
        return context


class ModifieProfesseur(EcoleAssignMixin, UpdateView):
    model = Proffeseur
    template_name = "cree_professeur.html"
    form_class = ProfesseurForm

    def form_valid(self, form):
        response = super().form_valid(form)  # self.object = prof MAJ
        prof = self.object

        tel = (prof.telephone or "").strip()
        created_prof_user = None

        # ✅ 1) Si user existe => synchroniser infos
        if prof.user_id:
            u = prof.user
            fields_to_update = []

            # adapte si tes champs User s'appellent différemment
            if hasattr(u, "nom_complet") and u.nom_complet != prof.nom_conplet:
                u.nom_complet = prof.nom_conplet
                fields_to_update.append("nom_complet")

            if hasattr(u, "num_tel") and tel and u.num_tel != tel:
                u.num_tel = tel
                fields_to_update.append("num_tel")

            # email seulement si prof.email est renseigné
            if getattr(prof, "email", None):
                prof_email = (prof.email or "").strip()
                if prof_email and u.email != prof_email:
                    u.email = prof_email
                    fields_to_update.append("email")

            if fields_to_update:
                u.save(update_fields=fields_to_update)

        # ✅ 2) Si pas de user et tel existe => créer user
        if not prof.user_id and tel:
            base = build_username(prof.nom_conplet, "prof")
            username = unique_username(User, base)
            email = (prof.email or "").strip() or f"{username}@prof.local"

            with transaction.atomic():
                created_prof_user = User.objects.create(
                    username=username,
                    email=email,
                    nom_complet=prof.nom_conplet,
                    num_tel=tel,
                    role="proffesseur",
                    ecole=prof.ecole,
                    password=make_password(tel)
                )
                prof.user = created_prof_user
                prof.save(update_fields=["user"])

        msg = _("Professeur modifié avec succès ✅")
        if created_prof_user:
            msg += f" | Prof login: {created_prof_user.username} | MDP: {tel}"
        messages.success(self.request, msg)

        return response

    def get_success_url(self):
        return reverse("Lesensiegnants")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["Title"] = _("Modifier un Professeur")
        context["Submit_text"] = _("Modifier")
        return context


class ProffesseurDelete(DeleteView):
    model = User
    template_name = 'DeleteProf.html'
    success_url = reverse_lazy('Lesensiegnants')



class ProffesseurMessagerie(ListView):
    model = Message
    template_name = 'MessageProf.html'







# views.py
import json
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST

from Ecole_admin.models import Violence
from Ecole_admin.models import Eleve, Classe, Niveau  # adapte si besoin


@login_required
def violence_list(request):
    q = (request.GET.get("q") or "").strip()

    violences = Violence.objects.select_related(
        "agresseur", "victime",
        "agresseur__classe", "victime__classe",
        "agresseur__classe__niveau", "victime__classe__niveau",
    )

    if q:
        violences = violences.filter(agresseur__nom__icontains=q) | violences.filter(victime__nom__icontains=q)

    violences = violences.order_by("-date")[:300]

    niveaux = Niveau.objects.all().order_by("nom")  # adapte "nom"
    return render(request, "violence_list.html", {
        "violences": violences,
        "niveaux": niveaux,
        "q": q,
    })


@login_required
def ajax_classes_by_niveau(request):
    niveau_id = request.GET.get("niveau_id")
    if not (niveau_id and niveau_id.isdigit()):
        return JsonResponse({"ok": True, "items": []})

    classes = Classe.objects.filter(niveau_id=int(niveau_id)).order_by("nom")
    items = [{"id": c.id, "text": c.nom} for c in classes]
    return JsonResponse({"ok": True, "items": items})


@login_required
def ajax_eleves_by_classe(request):
    classe_id = request.GET.get("classe_id")
    if not (classe_id and classe_id.isdigit()):
        return JsonResponse({"ok": True, "items": []})

    eleves = Eleve.objects.filter(classe_id=int(classe_id)).order_by("nom")
    items = [{"id": e.id, "text": getattr(e, "nom", str(e))} for e in eleves]
    return JsonResponse({"ok": True, "items": items})


@login_required
def violence_json(request, pk: int):
    v = get_object_or_404(
        Violence.objects.select_related(
            "agresseur__classe__niveau",
            "victime__classe__niveau",
        ),
        pk=pk
    )

    def pack_eleve(e):
        classe = getattr(e, "classe", None)
        niveau = getattr(classe, "niveau", None) if classe else None
        return {
            "eleve_id": e.id,
            "niveau_id": niveau.id if niveau else "",
            "classe_id": classe.id if classe else "",
        }

    return JsonResponse({
        "ok": True,
        "id": v.id,
        "agresseur": pack_eleve(v.agresseur),
        "victime": pack_eleve(v.victime),
        "forme_agression": v.forme_agression,
        "cause_violence": v.cause_violence,
        "dommage_corporel": v.dommage_corporel,
        "moyens_resolution": v.moyens_resolution,
        "date": v.date.strftime("%Y-%m-%d") if v.date else "",
    })


@login_required
@require_POST
def violence_save(request):
    """
    Reçoit JSON:
    {
      id?: number,
      ag_niveau_id, ag_classe_id, ag_eleve_id,
      vi_niveau_id, vi_classe_id, vi_eleve_id,
      forme_agression, cause_violence, dommage_corporel, moyens_resolution, date
    }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalide")

    vid = data.get("id")

    ag_eleve_id = str(data.get("ag_eleve_id") or "").strip()
    vi_eleve_id = str(data.get("vi_eleve_id") or "").strip()

    forme = (data.get("forme_agression") or "").strip()
    cause = (data.get("cause_violence") or "").strip()
    dommage = (data.get("dommage_corporel") or "").strip()
    moyens = (data.get("moyens_resolution") or "").strip()
    date = (data.get("date") or "").strip()

    if not (ag_eleve_id.isdigit() and vi_eleve_id.isdigit()):
        return JsonResponse({"ok": False, "message": "Champs manquants: agresseur et victime."}, status=400)

    # champs obligatoires comme ta capture (étoile rouge)
    if not (forme and cause and dommage and moyens and date):
        return JsonResponse({"ok": False, "message": "Veuillez remplir tous les champs obligatoires."}, status=400)

    ag = get_object_or_404(Eleve, pk=int(ag_eleve_id))
    vi = get_object_or_404(Eleve, pk=int(vi_eleve_id))

    if vid:
        v = get_object_or_404(Violence, pk=int(vid))
        v.agresseur = ag
        v.victime = vi
        v.forme_agression = forme
        v.cause_violence = cause
        v.dommage_corporel = dommage
        v.moyens_resolution = moyens
        v.date = date
        v.save()
        return JsonResponse({"ok": True, "message": "Modifié avec succès."})
    else:
        Violence.objects.create(
            agresseur=ag,
            victime=vi,
            forme_agression=forme,
            cause_violence=cause,
            dommage_corporel=dommage,
            moyens_resolution=moyens,
            date=date,
            cree_par=request.user
        )
        return JsonResponse({"ok": True, "message": "Ajouté avec succès."})


@login_required
@require_POST
def violence_delete(request, pk: int):
    v = get_object_or_404(Violence, pk=pk)
    v.delete()
    return JsonResponse({"ok": True, "message": "Supprimé."})





import io
import json
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.templatetags.static import static

from Ecole_admin.models import Degradation, Eleve, Classe, Niveau  # adapte si besoin

# ✅ Excel
import openpyxl
from openpyxl.styles import Font, Alignment

# ✅ WeasyPrint
from weasyprint import HTML, CSS


# -----------------------------
# LIST + AJAX + CRUD (inchangé)
# -----------------------------
@login_required
def degradation_page(request):
    q = (request.GET.get("q") or "").strip()

    qs = Degradation.objects.select_related("eleve", "eleve__classe")

    if q:
        qs = qs.filter(eleve__nom__icontains=q)

    degradations = qs.order_by("-date")[:300]
    niveaux = Niveau.objects.all().order_by("nom")

    return render(request, "degradation_list.html", {
        "degradations": degradations,
        "niveaux": niveaux,
        "q": q,
        "DECISIONS": Degradation.DECISIONS,
    })


@login_required
def ajax_classes_by_niveau(request):
    niveau_id = request.GET.get("niveau_id")
    if not (niveau_id and str(niveau_id).isdigit()):
        return JsonResponse({"ok": True, "items": []})

    classes = Classe.objects.filter(niveau_id=int(niveau_id)).order_by("nom")
    return JsonResponse({"ok": True, "items": [{"id": c.id, "text": c.nom} for c in classes]})


@login_required
def ajax_eleves_by_classe(request):
    classe_id = request.GET.get("classe_id")
    if not (classe_id and str(classe_id).isdigit()):
        return JsonResponse({"ok": True, "items": []})

    eleves = Eleve.objects.filter(classe_id=int(classe_id)).order_by("nom")
    return JsonResponse({"ok": True, "items": [{"id": e.id, "text": getattr(e, "nom", str(e))} for e in eleves]})


@login_required
def degradation_json(request, pk: int):
    d = get_object_or_404(Degradation.objects.select_related("eleve__classe__niveau"), pk=pk)
    eleve = d.eleve
    classe = getattr(eleve, "classe", None)
    niveau = getattr(classe, "niveau", None) if classe else None

    return JsonResponse({
        "ok": True,
        "id": d.id,
        "niveau_id": niveau.id if niveau else "",
        "classe_id": classe.id if classe else "",
        "eleve_id": eleve.id,
        "degradation_commise": d.degradation_commise,
        "decision_prise": d.decision_prise,
        "decision_autre": d.decision_autre or "",
        "date": d.date.strftime("%Y-%m-%d") if d.date else "",
    })


@login_required
@require_POST
def degradation_save(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("JSON invalide")

    pk = data.get("id")

    niveau_id = str(data.get("niveau_id") or "").strip()
    classe_id = str(data.get("classe_id") or "").strip()
    eleve_id = str(data.get("eleve_id") or "").strip()

    degradation_commise = (data.get("degradation_commise") or "").strip()
    decision_prise = (data.get("decision_prise") or "").strip()
    decision_autre = (data.get("decision_autre") or "").strip()
    date = (data.get("date") or "").strip()

    if not (niveau_id.isdigit() and classe_id.isdigit() and eleve_id.isdigit()):
        return JsonResponse({"ok": False, "message": "Veuillez sélectionner Niveau, Classe et Élève."}, status=400)

    if not (degradation_commise and decision_prise and date):
        return JsonResponse({"ok": False, "message": "Veuillez remplir tous les champs obligatoires."}, status=400)

    valid_decisions = {k for (k, _) in Degradation.DECISIONS}
    if decision_prise not in valid_decisions:
        return JsonResponse({"ok": False, "message": "Décision invalide."}, status=400)

    if decision_prise == "autre":
        if not decision_autre:
            return JsonResponse({"ok": False, "message": "Veuillez saisir l’autre décision."}, status=400)
    else:
        decision_autre = ""

    eleve = get_object_or_404(Eleve, pk=int(eleve_id))

    if pk:
        obj = get_object_or_404(Degradation, pk=int(pk))
        obj.eleve = eleve
        obj.degradation_commise = degradation_commise
        obj.decision_prise = decision_prise
        obj.decision_autre = decision_autre
        obj.date = date
        obj.save()
        return JsonResponse({"ok": True, "message": "Modification enregistrée."})
    else:
        Degradation.objects.create(
            eleve=eleve,
            degradation_commise=degradation_commise,
            decision_prise=decision_prise,
            decision_autre=decision_autre,
            date=date,
            cree_par=request.user
        )
        return JsonResponse({"ok": True, "message": "Ajout enregistré."})


@login_required
@require_POST
def degradation_delete(request, pk: int):
    obj = get_object_or_404(Degradation, pk=pk)
    obj.delete()
    return JsonResponse({"ok": True, "message": "Supprimé."})


import io
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.translation import gettext as _
from django.utils.translation import get_language

import openpyxl
from openpyxl.styles import Font, Alignment

from Ecole_admin.models import Degradation


@login_required
def degradation_export_excel(request):
    q = (request.GET.get("q") or "").strip()

    qs = Degradation.objects.select_related("eleve", "eleve__classe").order_by("-date")
    if q:
        qs = qs.filter(eleve__nom__icontains=q)

    # 🌍 langue active (fr / en / ar)
    lang = get_language() or "fr"

    wb = openpyxl.Workbook()
    ws = wb.active

    # 📄 Titre feuille (traduit)
    ws.title = _("Dégradations")

    # 🧾 Headers traduits
    headers = [
        _("Nom de l'élève"),
        _("Classe"),
        _("Dégradation commise"),
        _("Décision prise"),
        _("Date"),
    ]
    ws.append(headers)

    # style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # données
    for d in qs:
        eleve_nom = getattr(d.eleve, "nom", str(d.eleve))
        classe_nom = d.eleve.classe.nom if getattr(d.eleve, "classe", None) else ""

        ws.append([
            eleve_nom,
            classe_nom,
            d.degradation_commise,
            d.decision_affichee(),   # déjà traduit si tu utilises gettext dans le model
            d.date.strftime("%d-%m-%Y") if d.date else "",
        ])

    # largeur colonnes
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15

    # sortie
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 📦 nom fichier selon langue
    if lang.startswith("ar"):
        filename = "التخريبات.xlsx"
    elif lang.startswith("en"):
        filename = "damages.xlsx"
    else:
        filename = "degradations.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

# -----------------------------
# EXPORT PDF (WeasyPrint ✅)
# -----------------------------
@login_required
def degradation_export_pdf(request):
    q = (request.GET.get("q") or "").strip()

    qs = Degradation.objects.select_related("eleve", "eleve__classe", "eleve__classe__niveau").order_by("-date")
    if q:
        qs = qs.filter(eleve__nom__icontains=q)

    # ⚠️ Important: base_url pour que WeasyPrint charge /static/ et les images
    base_url = request.build_absolute_uri("/")

    # ✅ Chemin font (dans static/fonts/DejaVuSans.ttf)
    font_url = static("fonts/DejaVuSans.ttf")

    html_string = render_to_string("degradation_pdf.html", {
        "degradations": qs,
        "q": q,
        "font_url": font_url,  # utilisé dans le template
        "user": request.user,
    })

    pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf(
        stylesheets=[
            CSS(string="""
                @page { size: A4; margin: 14mm 12mm; }
            """)
        ]
    )

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="degradations.pdf"'
    return resp





import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template


import openpyxl
from openpyxl.styles import Font, Alignment

# PDF
from xhtml2pdf import pisa

from Ecole_admin.models import Violence   # ← ton modèle violence


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.utils.translation import gettext as _

# =========================
# EXPORT EXCEL – VIOLENCE (i18n FR / EN / AR)
# =========================
@login_required
def violence_export_excel(request):
    q = (request.GET.get("q") or "").strip()

    qs = Violence.objects.select_related(
        "agresseur",
        "victime",
        "agresseur__classe",
        "victime__classe",
    ).order_by("-date")

    if q:
        qs = qs.filter(Q(agresseur__nom__icontains=q) | Q(victime__nom__icontains=q))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _("Violences")

    # =========================
    # HEADERS (TRANSLATED)
    # =========================
    headers = [
        _("Agresseur"),
        _("Classe agresseur"),
        _("Victime"),
        _("Classe victime"),
        _("Forme d’agression"),
        _("Cause de la violence"),
        _("Dommage corporel"),
        _("Moyens de résolution"),
        _("Date"),
    ]
    ws.append(headers)

    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # =========================
    # DATA ROWS
    # =========================
    for v in qs:
        ws.append([
            v.agresseur.nom,
            v.agresseur.classe.nom if v.agresseur.classe else "",
            v.victime.nom,
            v.victime.classe.nom if v.victime.classe else "",
            v.forme_agression,
            v.cause_violence,
            v.dommage_corporel,
            v.moyens_resolution,
            v.date.strftime("%d-%m-%Y") if v.date else "",
        ])

    # =========================
    # COLUMN WIDTHS
    # =========================
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 32
    ws.column_dimensions["I"].width = 14

    # =========================
    # RESPONSE
    # =========================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="violences.xlsx"'
    return response




import io
import json
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.db.models import Q

from xhtml2pdf import pisa

from Ecole_admin.models import Niveau, Classe, Eleve, Violence  # ✅ adapte si ton model Violence est ailleurs


# -------------------- LANG (FR/EN/AR) --------------------
def _get_lang(request):
    lang = (request.GET.get("lang") or request.POST.get("lang") or "fr").strip().lower()
    return lang if lang in ("fr", "en", "ar") else "fr"

def _is_rtl(lang: str) -> bool:
    return lang == "ar"

def _t(lang: str):
    # ✅ mini dictionnaire (simple, fiable, pas besoin .po)
    FR = {
        "page_title": "Violences élèves",
        "title_list": "Violence pour les élèves",
        "btn_add": "Ajouter",
        "btn_excel": "Excel",
        "btn_pdf": "PDF",
        "btn_ok": "OK",
        "search_ph": "Rechercher ici",
        "th_agresseur": "L'agresseur (euse)",
        "th_forme": "Forme d’agression",
        "th_cause": "Cause de la violence",
        "th_dommage": "Dommage corporel",
        "th_resolution": "Moyens de résolution",
        "th_date": "Date",
        "th_actions": "Actions",
        "btn_edit_title": "Modifier",
        "btn_delete_title": "Supprimer",
        "empty": "Aucun enregistrement",
        "modal_title": "Enregistrement de violence",
        "lbl_niveau": "Niveaux",
        "lbl_classes": "La liste des classes",
        "lbl_agresseur": "L'agresseur (euse)",
        "lbl_forme": "Forme d'agression",
        "ph_forme": "Forme d'agression",
        "section_victime": "Victime",
        "lbl_victime": "Victime",
        "lbl_cause": "Cause de la violence",
        "ph_cause": "Cause de la violence",
        "lbl_dommage": "Dommage corporel",
        "ph_dommage": "Dommage corporel",
        "lbl_resolution": "Moyens de résolution",
        "ph_resolution": "Moyens de résolution",
        "lbl_date": "Date",
        "btn_close": "Fermer",
        "btn_save": "Sauvegarder",
        "select": "Select...",
        "pdf_title": "Violences",
        "pdf_header": "Rapport — Violences scolaires",
        "search_label": "Recherche",
        "th_classe": "Classe",
    }

    EN = {
        "page_title": "Student violence",
        "title_list": "Student violence records",
        "btn_add": "Add",
        "btn_excel": "Excel",
        "btn_pdf": "PDF",
        "btn_ok": "OK",
        "search_ph": "Search here",
        "th_agresseur": "Aggressor",
        "th_forme": "Type of aggression",
        "th_cause": "Cause",
        "th_dommage": "Injury / Damage",
        "th_resolution": "Resolution",
        "th_date": "Date",
        "th_actions": "Actions",
        "btn_edit_title": "Edit",
        "btn_delete_title": "Delete",
        "empty": "No records",
        "modal_title": "Violence record",
        "lbl_niveau": "Level",
        "lbl_classes": "Classes list",
        "lbl_agresseur": "Aggressor",
        "lbl_forme": "Type of aggression",
        "ph_forme": "Type of aggression",
        "section_victime": "Victim",
        "lbl_victime": "Victim",
        "lbl_cause": "Cause",
        "ph_cause": "Cause",
        "lbl_dommage": "Injury / Damage",
        "ph_dommage": "Injury / Damage",
        "lbl_resolution": "Resolution",
        "ph_resolution": "Resolution",
        "lbl_date": "Date",
        "btn_close": "Close",
        "btn_save": "Save",
        "select": "Select...",
        "pdf_title": "Violences",
        "pdf_header": "Report — School violence",
        "search_label": "Search",
        "th_classe": "Class",
    }

    AR = {
        "page_title": "عنف الطلاب",
        "title_list": "سجل عنف الطلاب",
        "btn_add": "إضافة",
        "btn_excel": "إكسل",
        "btn_pdf": "PDF",
        "btn_ok": "بحث",
        "search_ph": "ابحث هنا",
        "th_agresseur": "المعتدي",
        "th_forme": "نوع الاعتداء",
        "th_cause": "السبب",
        "th_dommage": "الضرر",
        "th_resolution": "الحل",
        "th_date": "التاريخ",
        "th_actions": "إجراءات",
        "btn_edit_title": "تعديل",
        "btn_delete_title": "حذف",
        "empty": "لا توجد بيانات",
        "modal_title": "تسجيل حالة عنف",
        "lbl_niveau": "المستوى",
        "lbl_classes": "قائمة الفصول",
        "lbl_agresseur": "المعتدي",
        "lbl_forme": "نوع الاعتداء",
        "ph_forme": "نوع الاعتداء",
        "section_victime": "الضحية",
        "lbl_victime": "الضحية",
        "lbl_cause": "السبب",
        "ph_cause": "السبب",
        "lbl_dommage": "الضرر",
        "ph_dommage": "الضرر",
        "lbl_resolution": "وسيلة الحل",
        "ph_resolution": "وسيلة الحل",
        "lbl_date": "التاريخ",
        "btn_close": "إغلاق",
        "btn_save": "حفظ",
        "select": "اختر...",
        "pdf_title": "العنف",
        "pdf_header": "تقرير — العنف المدرسي",
        "search_label": "بحث",
        "th_classe": "الفصل",
    }

    return {"fr": FR, "en": EN, "ar": AR}.get(lang, FR)

from django.contrib.staticfiles import finders
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily


def register_pdf_fonts():
    # Cherche le fichier dans static (peu importe l’emplacement réel)
    font_path = finders.find("fonts/DejaVuSans.ttf")
    if not font_path:
        raise Exception("DejaVuSans.ttf introuvable dans static/fonts/")

    # Enregistre la police (normal)
    try:
        pdfmetrics.getFont("DejaVuSans")
    except Exception:
        pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))

    # ✅ Si tu n’as pas Bold, on mappe le bold sur la même police
    registerFontFamily("DejaVuSans", normal="DejaVuSans", bold="DejaVuSans")



import re
import arabic_reshaper
from bidi.algorithm import get_display

_arabic_re = re.compile(r'[\u0600-\u06FF]')

def shape_arabic(text: str) -> str:
    if not text:
        return text
    # si pas d’arabe, retourne tel quel
    if not _arabic_re.search(text):
        return text
    # reshape + bidi
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)



import os
from django.conf import settings
from django.contrib.staticfiles import finders

def link_callback(uri, rel):
    """
    Convertit les URIs HTML (static/media) en chemins absolus
    pour xhtml2pdf, sans sortir des dossiers autorisés.
    """

    # 1) STATIC_URL -> chemin relatif pour finders
    if uri.startswith(settings.STATIC_URL):
        relative_path = uri.replace(settings.STATIC_URL, "", 1)
        absolute_path = finders.find(relative_path)
        if absolute_path:
            # finders.find peut retourner list
            if isinstance(absolute_path, (list, tuple)):
                absolute_path = absolute_path[0]
            return absolute_path

        # fallback STATIC_ROOT si collectstatic
        if getattr(settings, "STATIC_ROOT", None):
            absolute_path = os.path.join(settings.STATIC_ROOT, relative_path)
            if os.path.isfile(absolute_path):
                return absolute_path

    # 2) MEDIA_URL
    if uri.startswith(settings.MEDIA_URL):
        relative_path = uri.replace(settings.MEDIA_URL, "", 1)
        absolute_path = os.path.join(settings.MEDIA_ROOT, relative_path)
        if os.path.isfile(absolute_path):
            return absolute_path

    # 3) deja un chemin local absolu
    if os.path.isfile(uri):
        return uri

    raise Exception(f"Impossible de trouver la ressource: {uri}")

# -------------------- PDF helper --------------------
def _render_pdf(template_src, context):
    register_pdf_fonts()  # ✅ IMPORTANT (fix le temp .ttf)

    template = get_template(template_src)
    html = template.render(context)

    result = io.BytesIO()
    pdf = pisa.pisaDocument(
        io.BytesIO(html.encode("utf-8")),
        result,
        encoding="utf-8",
        # link_callback=link_callback  # optionnel si tu as images
    )
    if pdf.err:
        return None
    return result.getvalue()




# -------------------- Pages --------------------
@login_required
def violence_list(request):
    lang = _get_lang(request)
    t = _t(lang)
    is_rtl = _is_rtl(lang)

    q = (request.GET.get("q") or "").strip()

    qs = Violence.objects.select_related(
        "agresseur", "victime", "agresseur__classe", "victime__classe"
    ).order_by("-date")

    if q:
        qs = qs.filter(Q(agresseur__nom__icontains=q) | Q(victime__nom__icontains=q))

    niveaux = Niveau.objects.all()
    if hasattr(Niveau, "actif"):
        niveaux = niveaux.filter(actif=True)

    js_i18n = {
        "select": t["select"],
        "confirm_delete": {"fr": "Supprimer cet enregistrement ?", "en": "Delete this record?", "ar": "هل تريد حذف هذا السجل؟"}[lang],
        "err_load": {"fr": "Erreur chargement.", "en": "Loading error.", "ar": "خطأ في التحميل."}[lang],
        "err_delete": {"fr": "Erreur suppression", "en": "Delete error", "ar": "خطأ في الحذف"}[lang],
        "err_save": {"fr": "Erreur sauvegarde.", "en": "Save error.", "ar": "خطأ في الحفظ."}[lang],
        "ok_saved": {"fr": "Enregistré.", "en": "Saved.", "ar": "تم الحفظ."}[lang],
        "err_network": {"fr": "Erreur réseau.", "en": "Network error.", "ar": "خطأ في الشبكة."}[lang],
    }

    from django.shortcuts import render
    return render(request, "violence_list.html", {
        "violences": qs,
        "q": q,
        "niveaux": niveaux,
        "lang": lang,
        "is_rtl": is_rtl,
        "t": t,
        "js_i18n": json.dumps(js_i18n, ensure_ascii=False),
    })


import re
import arabic_reshaper
from bidi.algorithm import get_display
from django.utils.translation import get_language

_AR_RE = re.compile(r"[\u0600-\u06FF]")

def pdf_text(s: str) -> str:
    """
    Fix RTL + shaping pour l’arabe dans xhtml2pdf.
    Si la langue active n’est pas 'ar' -> retourne tel quel.
    """
    if not s:
        return ""
    lang = (get_language() or "").lower()
    if not lang.startswith("ar"):
        return s

    if not _AR_RE.search(s):
        return s

    reshaped = arabic_reshaper.reshape(s)
    return get_display(reshaped)

import os
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from django.template.loader import get_template
from django.contrib.staticfiles import finders
from django.utils.translation import get_language

from weasyprint import HTML

# adapte l'import de ton model Violence
from Ecole_admin.models import Violence

@login_required
def violence_export_pdf(request):
    q = (request.GET.get("q") or "").strip()

    qs = Violence.objects.select_related(
        "agresseur",
        "victime",
        "agresseur__classe",
        "victime__classe",
    ).order_by("-date")

    if q:
        qs = qs.filter(Q(agresseur__nom__icontains=q) | Q(victime__nom__icontains=q))

    # ✅ chemin absolu de la police (dans static/fonts)
    font_path = finders.find("fonts/DejaVuSans.ttf")
    if not font_path:
        return HttpResponse("Font DejaVuSans.ttf introuvable dans static/fonts", status=500)

    # rendu HTML depuis template
    html_string = get_template("violence_pdf.html").render({
        "violences": qs,
        "q": q,
        "font_path": font_path,
    })

    # ✅ base_url IMPORTANT: permet de résoudre static/media si tu en ajoutes plus tard
    base_url = request.build_absolute_uri("/")

    pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="violences.pdf"'
    return response



# -------------------- AJAX: classes by niveau --------------------
@login_required
def ajax_classes_by_niveau_violence(request):
    niveau_id = request.GET.get("niveau_id")
    if not niveau_id:
        return JsonResponse({"ok": True, "items": []})

    qs = Classe.objects.filter(niveau_id=niveau_id)
    if hasattr(Classe, "actif"):
        qs = qs.filter(actif=True)

    items = [{"id": c.id, "text": c.nom} for c in qs.order_by("nom")]
    return JsonResponse({"ok": True, "items": items})


# -------------------- AJAX: eleves by classe --------------------
@login_required
def ajax_eleves_by_classe_violence(request):
    classe_id = request.GET.get("classe_id")
    if not classe_id:
        return JsonResponse({"ok": True, "items": []})

    qs = Eleve.objects.filter(classe_id=classe_id).order_by("nom")
    items = [{"id": e.id, "text": e.nom} for e in qs]
    return JsonResponse({"ok": True, "items": items})


# -------------------- JSON record (for edit) --------------------
@login_required
def violence_json(request, pk):
    try:
        v = Violence.objects.select_related(
            "agresseur", "victime", "agresseur__classe", "victime__classe",
            "agresseur__classe__niveau", "victime__classe__niveau",
        ).get(pk=pk)
    except Violence.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Not found"}, status=404)

    def pack_eleve(e):
        return {
            "eleve_id": e.id if e else None,
            "classe_id": e.classe_id if e else None,
            "niveau_id": (e.classe.niveau_id if (e and e.classe_id and hasattr(e.classe, "niveau_id")) else None),
        }

    data = {
        "ok": True,
        "id": v.id,
        "agresseur": pack_eleve(v.agresseur),
        "victime": pack_eleve(v.victime),
        "forme_agression": v.forme_agression,
        "cause_violence": v.cause_violence,
        "dommage_corporel": v.dommage_corporel,
        "moyens_resolution": v.moyens_resolution,
        "date": v.date.isoformat() if v.date else "",
    }
    return JsonResponse(data)


# -------------------- Save (create/update) --------------------
@login_required
def violence_save(request):
    lang = _get_lang(request)
    t = _t(lang)

    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Method not allowed"}, status=405)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "message": "Invalid JSON"}, status=400)

    vid = payload.get("id")
    ag_id = payload.get("ag_eleve_id")
    vi_id = payload.get("vi_eleve_id")

    forme = (payload.get("forme_agression") or "").strip()
    cause = (payload.get("cause_violence") or "").strip()
    dommage = (payload.get("dommage_corporel") or "").strip()
    resolution = (payload.get("moyens_resolution") or "").strip()
    date = payload.get("date")

    # validation
    if not (ag_id and vi_id and forme and cause and dommage and resolution and date):
        msg = {"fr": "Champs manquants.", "en": "Missing fields.", "ar": "حقول ناقصة."}.get(lang, "Champs manquants.")
        return JsonResponse({"ok": False, "message": msg}, status=400)

    if str(ag_id) == str(vi_id):
        msg = {"fr": "Agresseur et victime ne peuvent pas être la même personne.",
               "en": "Aggressor and victim cannot be the same person.",
               "ar": "لا يمكن أن يكون المعتدي والضحية نفس الشخص."}.get(lang)
        return JsonResponse({"ok": False, "message": msg}, status=400)

    try:
        ag = Eleve.objects.get(pk=ag_id)
        vi = Eleve.objects.get(pk=vi_id)
    except Eleve.DoesNotExist:
        msg = {"fr": "Élève introuvable.", "en": "Student not found.", "ar": "الطالب غير موجود."}.get(lang)
        return JsonResponse({"ok": False, "message": msg}, status=404)

    if vid:
        try:
            v = Violence.objects.get(pk=vid)
        except Violence.DoesNotExist:
            return JsonResponse({"ok": False, "message": "Not found"}, status=404)
    else:
        v = Violence()

    v.agresseur = ag
    v.victime = vi
    v.forme_agression = forme
    v.cause_violence = cause
    v.dommage_corporel = dommage
    v.moyens_resolution = resolution
    v.date = date  # Django convertit ISO YYYY-MM-DD

    v.save()

    okmsg = {"fr": "Enregistrement sauvegardé.",
             "en": "Record saved.",
             "ar": "تم حفظ السجل."}.get(lang)
    return JsonResponse({"ok": True, "message": okmsg})


# -------------------- Delete --------------------
@login_required
def violence_delete(request, pk):
    lang = _get_lang(request)

    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Method not allowed"}, status=405)

    try:
        v = Violence.objects.get(pk=pk)
    except Violence.DoesNotExist:
        msg = {"fr": "Introuvable.", "en": "Not found.", "ar": "غير موجود."}.get(lang)
        return JsonResponse({"ok": False, "message": msg}, status=404)

    v.delete()
    msg = {"fr": "Supprimé.", "en": "Deleted.", "ar": "تم الحذف."}.get(lang)
    return JsonResponse({"ok": True, "message": msg})






