from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404

from Ecole_admin.models import ProgrammeOrientation
from Ecole_admin.form import ProgrammeOrientationForm


def get_ecole_user(request):
    return request.user.ecole


@login_required
def programme_orientation_list(request):
    ecole = get_ecole_user(request)

    items = ProgrammeOrientation.objects.filter(
        ecole=ecole
    ).order_by("ordre", "code")

    context = {
        "items": items,
        "page_title": "Programmes d'orientation",
    }
    return render(request, "orientation/programme_orientation_list.html", context)


@login_required
def programme_orientation_create(request):
    ecole = get_ecole_user(request)

    if request.method == "POST":
        form = ProgrammeOrientationForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.save()
            messages.success(request, "Le programme d'orientation a été ajouté avec succès.")
            return redirect("programme_orientation_list")
    else:
        form = ProgrammeOrientationForm()

    context = {
        "form": form,
        "titre": "Ajouter un programme d'orientation",
    }
    return render(request, "orientation/programme_orientation_form.html", context)


@login_required
def programme_orientation_update(request, pk):
    ecole = get_ecole_user(request)
    obj = get_object_or_404(ProgrammeOrientation, pk=pk, ecole=ecole)

    if request.method == "POST":
        form = ProgrammeOrientationForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Le programme d'orientation a été modifié avec succès.")
            return redirect("programme_orientation_list")
    else:
        form = ProgrammeOrientationForm(instance=obj)

    context = {
        "form": form,
        "titre": "Modifier un programme d'orientation",
        "obj": obj,
    }
    return render(request, "orientation/programme_orientation_form.html", context)


@login_required
def programme_orientation_delete(request, pk):
    ecole = get_ecole_user(request)
    obj = get_object_or_404(ProgrammeOrientation, pk=pk, ecole=ecole)

    if request.method == "POST":
        obj.delete()
        messages.success(request, "Le programme d'orientation a été supprimé avec succès.")
        return redirect("programme_orientation_list")

    context = {
        "obj": obj,
    }
    return render(request, "orientation/programme_orientation_delete.html", context)




from Ecole_admin.models import OrientationScolaire, AnneeScolaire
from Ecole_admin.form import OrientationScolaireForm


@login_required
def orientation_scolaire_list(request):
    ecole = get_ecole_user(request)
    annee_active = AnneeScolaire.objects.filter(est_active=True).first()

    items = OrientationScolaire.objects.filter(
        ecole=ecole,
        annee_scolaire=annee_active
    ).select_related("niveau").prefetch_related("programmes").order_by("niveau__ordre", "niveau__nom")

    context = {
        "items": items,
        "annee_active": annee_active,
        "page_title": "Orientation scolaire",
    }
    return render(request, "orientation/orientation_scolaire_list.html", context)


@login_required
def orientation_scolaire_create(request):
    ecole = get_ecole_user(request)
    annee_active = AnneeScolaire.objects.filter(est_active=True).first()

    if request.method == "POST":
        form = OrientationScolaireForm(request.POST, ecole=ecole)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.annee_scolaire = annee_active
            obj.save()
            form.save_m2m()
            messages.success(request, "L'orientation scolaire a été ajoutée avec succès.")
            return redirect("orientation_scolaire_list")
    else:
        form = OrientationScolaireForm(ecole=ecole)

    context = {
        "form": form,
        "titre": "Ajouter une orientation scolaire",
    }
    return render(request, "orientation/orientation_scolaire_form.html", context)


@login_required
def orientation_scolaire_update(request, pk):
    ecole = get_ecole_user(request)
    obj = get_object_or_404(OrientationScolaire, pk=pk, ecole=ecole)

    if request.method == "POST":
        form = OrientationScolaireForm(request.POST, instance=obj, ecole=ecole)
        if form.is_valid():
            form.save()
            messages.success(request, "L'orientation scolaire a été modifiée avec succès.")
            return redirect("orientation_scolaire_list")
    else:
        form = OrientationScolaireForm(instance=obj, ecole=ecole)

    context = {
        "form": form,
        "titre": "Modifier une orientation scolaire",
        "obj": obj,
    }
    return render(request, "orientation/orientation_scolaire_form.html", context)


@login_required
def orientation_scolaire_delete(request, pk):
    ecole = get_ecole_user(request)
    obj = get_object_or_404(OrientationScolaire, pk=pk, ecole=ecole)

    if request.method == "POST":
        obj.delete()
        messages.success(request, "L'orientation scolaire a été supprimée avec succès.")
        return redirect("orientation_scolaire_list")

    context = {
        "obj": obj,
    }
    return render(request, "orientation/orientation_scolaire_delete.html", context)





from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from Ecole_admin.form import FicheVoeuxFilterForm
from django.http import JsonResponse

from Ecole_admin.models import (
    Eleve,
    OrientationScolaire,
    AnneeScolaire,
    Niveau,
    Classe,
)



def get_ecole_user(request):
    return request.user.ecole


@login_required
def fiche_voeux_list(request):
    ecole = get_ecole_user(request)
    form = FicheVoeuxFilterForm(request.GET or None, ecole=ecole)

    eleves = []
    orientation = None
    classe = None

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).prefetch_related("programmes").first()

        if orientation:
            eleves = Eleve.objects.filter(
                ecole=ecole,
                annee_scolaire=annee_scolaire,
                classe=classe
            ).order_by("nom")

    context = {
        "form": form,
        "eleves": eleves,
        "orientation": orientation,
        "classe": classe,
    }
    return render(request, "orientation/fiche_voeux_list.html", context)


@login_required
def ajax_niveaux_orientation(request):
    ecole = get_ecole_user(request)
    annee_id = request.GET.get("annee_scolaire")

    data = []

    if annee_id:
        niveau_ids = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire_id=annee_id,
            actif=True
        ).values_list("niveau_id", flat=True)

        niveaux = Niveau.objects.filter(
            ecole=ecole,
            actif=True,
            id__in=niveau_ids
        ).order_by("ordre", "nom")

        data = [{"id": n.id, "nom": n.nom} for n in niveaux]

    return JsonResponse({"results": data})


@login_required
def ajax_classes_par_niveau(request):
    ecole = get_ecole_user(request)
    niveau_id = request.GET.get("niveau")

    data = []

    if niveau_id:
        classes = Classe.objects.filter(
            ecole=ecole,
            actif=True,
            niveau_id=niveau_id
        ).order_by("nom")

        data = [{"id": c.id, "nom": c.nom} for c in classes]

    return JsonResponse({"results": data})



from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render, redirect

from Ecole_admin.models import (
    OrientationScolaire,
    Niveau,
    Classe,
    Eleve,
    VoeuOrientationEleve,
    VoeuOrientationChoix,
)
from Ecole_admin.form import SaisieVoeuxFilterForm


def get_ecole_user(request):
    return request.user.ecole


@login_required
def saisie_voeux_orientation(request):
    ecole = get_ecole_user(request)
    data_source = request.POST if request.method == "POST" else request.GET
    form = SaisieVoeuxFilterForm(data_source or None, ecole=ecole)

    rows = []
    orientation = None
    programmes = []
    classe = None
    annee_scolaire = None
    niveau = None

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).prefetch_related("programmes").first()

        if orientation:
            programmes = list(orientation.programmes.all().order_by("ordre", "code"))

            eleves = list(
                Eleve.objects.filter(
                    ecole=ecole,
                    annee_scolaire=annee_scolaire,
                    classe=classe
                ).order_by("nom")
            )

            if request.method == "POST":
                for eleve in eleves:
                    voeu, _ = VoeuOrientationEleve.objects.get_or_create(
                        eleve=eleve,
                        orientation=orientation,
                        defaults={"cree_par": request.user}
                    )

                    voeu.choix.all().delete()

                    selected_count = 0
                    for programme in programmes:
                        field_name = f"choix_{eleve.id}_{programme.id}"
                        if request.POST.get(field_name):
                            VoeuOrientationChoix.objects.create(
                                voeu=voeu,
                                programme=programme
                            )
                            selected_count += 1

                    if selected_count == 0:
                        voeu.delete()

                messages.success(request, "Les choix d'orientation ont été enregistrés avec succès.")
                return redirect(
                    f"{request.path}?annee_scolaire={annee_scolaire.id}&niveau={niveau.id}&classe={classe.id}"
                )

            existing = VoeuOrientationEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).prefetch_related("choix__programme")

            existing_map = {}
            for item in existing:
                existing_map[item.eleve_id] = [c.programme_id for c in item.choix.all()]

            for eleve in eleves:
                rows.append({
                    "eleve": eleve,
                    "selected_program_ids": existing_map.get(eleve.id, [])
                })

    context = {
        "form": form,
        "rows": rows,
        "orientation": orientation,
        "programmes": programmes,
        "classe": classe,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
    }
    return render(request, "orientation/saisie_voeux_orientation.html", context)


@login_required
def ajax_niveaux_orientation(request):
    ecole = get_ecole_user(request)
    annee_id = request.GET.get("annee_scolaire")

    results = []

    if annee_id:
        niveau_ids = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire_id=annee_id,
            actif=True
        ).values_list("niveau_id", flat=True)

        niveaux = Niveau.objects.filter(
            ecole=ecole,
            actif=True,
            id__in=niveau_ids
        ).order_by("ordre", "nom")

        results = [{"id": n.id, "nom": n.nom} for n in niveaux]

    return JsonResponse({"results": results})


@login_required
def ajax_classes_par_niveau(request):
    ecole = get_ecole_user(request)
    niveau_id = request.GET.get("niveau")

    results = []

    if niveau_id:
        classes = Classe.objects.filter(
            ecole=ecole,
            actif=True,
            niveau_id=niveau_id
        ).order_by("nom")

        results = [{"id": c.id, "nom": c.nom} for c in classes]

    return JsonResponse({"results": results})



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from Ecole_admin.models import (
    OrientationScolaire,
    Eleve,
    VoeuOrientationEleve,
)
from Ecole_admin.form import AffichageVoeuxFilterForm


def get_ecole_user(request):
    return request.user.ecole


@login_required
def voeux_eleves_list(request):
    ecole = get_ecole_user(request)
    form = AffichageVoeuxFilterForm(request.GET or None, ecole=ecole)

    rows = []
    orientation = None
    annee_scolaire = None
    niveau = None
    classe = None

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).prefetch_related("programmes").first()

        if orientation:
            eleves = Eleve.objects.filter(
                ecole=ecole,
                annee_scolaire=annee_scolaire,
                classe=classe
            ).order_by("nom")

            voeux = VoeuOrientationEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).prefetch_related("choix__programme")

            voeux_map = {v.eleve_id: v for v in voeux}

            for eleve in eleves:
                voeu = voeux_map.get(eleve.id)
                choix_codes = []

                if voeu:
                    choix_codes = [
                        item.programme.code
                        for item in voeu.choix.all().order_by("programme__ordre", "programme__code")
                    ]

                rows.append({
                    "eleve": eleve,
                    "voeu": voeu,
                    "choix_codes": choix_codes,
                    "a_voeux": bool(voeu and choix_codes),
                })

    context = {
        "form": form,
        "rows": rows,
        "orientation": orientation,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
        "classe": classe,
    }
    return render(request, "orientation/voeux_eleves_list.html", context)


@login_required
def voeu_eleve_detail(request, eleve_id, orientation_id):
    ecole = get_ecole_user(request)

    eleve = get_object_or_404(Eleve, pk=eleve_id, ecole=ecole)
    orientation = get_object_or_404(
        OrientationScolaire,
        pk=orientation_id,
        ecole=ecole
    )

    voeu = VoeuOrientationEleve.objects.filter(
        eleve=eleve,
        orientation=orientation
    ).prefetch_related("choix__programme").first()

    selected_ids = []
    if voeu:
        selected_ids = [item.programme_id for item in voeu.choix.all()]

    programmes = orientation.programmes.all().order_by("ordre", "code")

    context = {
        "eleve": eleve,
        "orientation": orientation,
        "voeu": voeu,
        "selected_ids": selected_ids,
        "programmes": programmes,
    }
    return render(request, "orientation/voeu_eleve_detail.html", context)



from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from Ecole_admin.models import (
    OrientationScolaire,
    PeriodeScolaire,
    Eleve,
    Note,
    VoeuOrientationEleve,
    OrientationFinaleEleve,
)
from Ecole_admin.form import OrientationFinaleFilterForm


def get_ecole_user(request):
    return request.user.ecole


from decimal import Decimal


def moyenne_generale_par_periode(eleve, annee_scolaire, periode, ordre_periode):
    notes_qs = Note.objects.filter(
        eleve=eleve,
        annee_scolaire=annee_scolaire,
    ).select_related("devoir")

    notes_par_devoir = notes_qs.filter(
        devoir__isnull=False,
        devoir__periode=periode
    )

    if notes_par_devoir.exists():
        notes = list(notes_par_devoir)
    else:
        notes = list(notes_qs.filter(trimestre=ordre_periode))

    if not notes:
        return None

    total = Decimal("0")
    count = 0

    for note in notes:
        total += Decimal(str(note.note or 0))
        count += 1

    if count == 0:
        return None

    return round(total / Decimal(count), 2)


from decimal import Decimal


def moyenne_annuelle_depuis_trimestres(eleve, annee_scolaire, periodes):
    """
    Logique temporaire validée :
    - chaque trimestre compte dans l'année
    - même si un trimestre n'a pas de note, il compte comme 0
    - la division se fait par le nombre total de trimestres enregistrés
    """

    if not periodes:
        return None

    total = Decimal("0")

    for idx, periode in enumerate(periodes, start=1):
        moyenne = moyenne_generale_par_periode(
            eleve=eleve,
            annee_scolaire=annee_scolaire,
            periode=periode,
            ordre_periode=idx
        )

        if moyenne is not None:
            total += Decimal(str(moyenne))
        else:
            total += Decimal("0")

    return round(total / Decimal(len(periodes)), 2)


@login_required
def orientation_finale_page(request):
    ecole = get_ecole_user(request)
    data_source = request.POST if request.method == "POST" else request.GET
    form = OrientationFinaleFilterForm(data_source or None, ecole=ecole)

    rows = []
    orientation = None
    periodes = []
    programmes = []
    annee_scolaire = None
    niveau = None
    classe = None

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).prefetch_related("programmes").first()

        if orientation:
            periodes = list(
                PeriodeScolaire.objects.filter(
                    ecole=ecole,
                    annee_scolaire=annee_scolaire
                ).order_by("debut")
            )
            programmes = list(orientation.programmes.all().order_by("ordre", "code"))

            eleves = list(
                Eleve.objects.filter(
                    ecole=ecole,
                    annee_scolaire=annee_scolaire,
                    classe=classe
                ).order_by("nom")
            )

            if request.method == "POST":
                for eleve in eleves:
                    programme_id = request.POST.get(f"programme_final_{eleve.id}")
                    remarque = (request.POST.get(f"remarque_{eleve.id}") or "").strip()

                    if programme_id:
                        programme = next((p for p in programmes if str(p.id) == str(programme_id)), None)
                        if programme:
                            OrientationFinaleEleve.objects.update_or_create(
                                eleve=eleve,
                                orientation=orientation,
                                defaults={
                                    "programme_final": programme,
                                    "remarque": remarque,
                                    "cree_par": request.user,
                                }
                            )
                    else:
                        OrientationFinaleEleve.objects.filter(
                            eleve=eleve,
                            orientation=orientation
                        ).delete()

                messages.success(request, "Les orientations finales ont été enregistrées avec succès.")
                return redirect(
                    f"{request.path}?annee_scolaire={annee_scolaire.id}&niveau={niveau.id}&classe={classe.id}"
                )

            voeux_qs = VoeuOrientationEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).prefetch_related("choix__programme")

            voeux_map = {item.eleve_id: item for item in voeux_qs}

            finales_qs = OrientationFinaleEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).select_related("programme_final")

            finales_map = {item.eleve_id: item for item in finales_qs}

            for eleve in eleves:
                moyennes = []

                for idx, periode in enumerate(periodes, start=1):
                    moyenne = moyenne_generale_par_periode(
                        eleve=eleve,
                        annee_scolaire=annee_scolaire,
                        periode=periode,
                        ordre_periode=idx
                    )
                    moyennes.append({
                        "periode": periode,
                        "moyenne": moyenne,
                    })

                moyenne_annuelle = moyenne_annuelle_depuis_trimestres(
                    eleve=eleve,
                    annee_scolaire=annee_scolaire,
                    periodes=periodes
                )

                voeu = voeux_map.get(eleve.id)
                choix_codes = []
                if voeu:
                    choix_codes = [
                        item.programme.code
                        for item in voeu.choix.all().order_by("programme__ordre", "programme__code")
                    ]

                finale = finales_map.get(eleve.id)

                rows.append({
                    "eleve": eleve,
                    "moyennes": moyennes,
                    "moyenne_annuelle": moyenne_annuelle,
                    "choix_codes": choix_codes,
                    "finale": finale,
                })

    context = {
        "form": form,
        "rows": rows,
        "orientation": orientation,
        "periodes": periodes,
        "programmes": programmes,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
        "classe": classe,
    }
    return render(request, "orientation/orientation_finale_page.html", context)





from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from Ecole_admin.models import (
    OrientationScolaire,
    OrientationFinaleEleve,
    Classe,
    Eleve,
)
from Ecole_admin.form import RapportOrientationFilterForm


def get_ecole_user(request):
    return request.user.ecole


@login_required
def rapport_orientation_page(request):
    ecole = get_ecole_user(request)
    form = RapportOrientationFilterForm(request.GET or None, ecole=ecole)

    orientation = None
    annee_scolaire = None
    niveau = None
    classes = []
    programmes = []
    matrix = []
    class_totals = []
    total_general = 0
    total_eleves = 0
    total_orientes = 0
    total_non_orientes = 0

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).prefetch_related("programmes").first()

        classes = list(
            Classe.objects.filter(
                ecole=ecole,
                actif=True,
                niveau=niveau
            ).order_by("nom")
        )

        eleves_qs = Eleve.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            classe__in=classes
        )

        total_eleves = eleves_qs.count()

        if orientation:
            programmes = list(orientation.programmes.all().order_by("ordre", "code"))

            finales_qs = OrientationFinaleEleve.objects.filter(
                orientation=orientation,
                eleve__classe__in=classes,
                eleve__annee_scolaire=annee_scolaire
            ).select_related("eleve__classe", "programme_final")

            total_orientes = finales_qs.count()
            total_non_orientes = max(total_eleves - total_orientes, 0)

            for programme in programmes:
                row_total = 0
                class_counts = []

                for classe in classes:
                    count = finales_qs.filter(
                        eleve__classe=classe,
                        programme_final=programme
                    ).count()

                    class_counts.append({
                        "classe": classe,
                        "count": count,
                    })
                    row_total += count

                total_general += row_total

                matrix.append({
                    "programme": programme,
                    "class_counts": class_counts,
                    "total": row_total,
                })

            for classe in classes:
                count = finales_qs.filter(
                    eleve__classe=classe
                ).count()

                class_totals.append({
                    "classe": classe,
                    "count": count,
                })

    context = {
        "form": form,
        "orientation": orientation,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
        "classes": classes,
        "programmes": programmes,
        "matrix": matrix,
        "class_totals": class_totals,
        "total_general": total_general,
        "total_eleves": total_eleves,
        "total_orientes": total_orientes,
        "total_non_orientes": total_non_orientes,
    }
    return render(request, "orientation/rapport_orientation_page.html", context)





from decimal import Decimal
from Ecole_admin.models import Note


def moyenne_generale_par_periode(eleve, annee_scolaire, periode, ordre_periode):
    notes_qs = Note.objects.filter(
        eleve=eleve,
        annee_scolaire=annee_scolaire,
    ).select_related("devoir")

    notes_par_devoir = notes_qs.filter(
        devoir__isnull=False,
        devoir__periode=periode
    )

    if notes_par_devoir.exists():
        notes = list(notes_par_devoir)
    else:
        notes = list(notes_qs.filter(trimestre=ordre_periode))

    if not notes:
        return None

    total = Decimal("0")
    count = 0

    for note in notes:
        total += Decimal(str(note.note or 0))
        count += 1

    if count == 0:
        return None

    return round(total / Decimal(count), 2)


def moyenne_annuelle_depuis_trimestres(eleve, annee_scolaire, periodes):
    if not periodes:
        return None

    total = Decimal("0")

    for idx, periode in enumerate(periodes, start=1):
        moyenne = moyenne_generale_par_periode(
            eleve=eleve,
            annee_scolaire=annee_scolaire,
            periode=periode,
            ordre_periode=idx
        )
        if moyenne is not None:
            total += Decimal(str(moyenne))
        else:
            total += Decimal("0")

    return round(total / Decimal(len(periodes)), 2)




from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from Ecole_admin.models import (
    OrientationScolaire,
    Eleve,
    OrientationFinaleEleve,
    DecisionFinaleEleve,
)
from Ecole_admin.form import DecisionEleveFilterForm



def get_ecole_user(request):
    return request.user.ecole


@login_required
def decision_eleve_page(request):
    ecole = get_ecole_user(request)
    form = DecisionEleveFilterForm(request.GET or None, ecole=ecole)

    rows = []
    orientation = None
    annee_scolaire = None
    niveau = None
    classe = None

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation = OrientationScolaire.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            actif=True
        ).first()

        if orientation:
            eleves = list(
                Eleve.objects.filter(
                    ecole=ecole,
                    annee_scolaire=annee_scolaire,
                    classe=classe
                ).select_related("classe").order_by("nom")
            )

            # choix de l'élève
            voeux_qs = VoeuOrientationEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).prefetch_related("choix__programme")

            voeux_map = {item.eleve_id: item for item in voeux_qs}

            # décision de l'administration
            finales_qs = OrientationFinaleEleve.objects.filter(
                eleve__in=eleves,
                orientation=orientation
            ).select_related("programme_final")

            finales_map = {item.eleve_id: item for item in finales_qs}

            for eleve in eleves:
                voeu = voeux_map.get(eleve.id)
                finale = finales_map.get(eleve.id)

                choix_codes = []
                if voeu:
                    choix_codes = [
                        item.programme.code
                        for item in voeu.choix.all().order_by("programme__ordre", "programme__code")
                    ]

                rows.append({
                    "eleve": eleve,
                    "choix_eleve": ", ".join(choix_codes) if choix_codes else "",
                    "decision_administration": finale.programme_final.code if finale and finale.programme_final else "",
                })

    context = {
        "form": form,
        "rows": rows,
        "orientation": orientation,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
        "classe": classe,
    }
    return render(request, "orientation/decision_eleve_page.html", context)





from io import BytesIO
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from Ecole_admin.models import (
    OrientationScolaire,
    Eleve,
    OrientationFinaleEleve,
    VoeuOrientationEleve,
)
from Ecole_admin.form import RapportDecisionFinaleFilterForm


def get_ecole_user(request):
    return request.user.ecole


def build_rapport_decision_finale_rows(ecole, annee_scolaire, niveau, classe):
    orientation = OrientationScolaire.objects.filter(
        ecole=ecole,
        annee_scolaire=annee_scolaire,
        niveau=niveau,
        actif=True
    ).first()

    rows = []

    if not orientation:
        return orientation, rows

    eleves = list(
        Eleve.objects.filter(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            classe=classe
        ).select_related("classe").order_by("nom")
    )

    voeux_qs = VoeuOrientationEleve.objects.filter(
        eleve__in=eleves,
        orientation=orientation
    ).prefetch_related("choix__programme")

    voeux_map = {item.eleve_id: item for item in voeux_qs}

    finales_qs = OrientationFinaleEleve.objects.filter(
        eleve__in=eleves,
        orientation=orientation
    ).select_related("programme_final")

    finales_map = {item.eleve_id: item for item in finales_qs}

    for eleve in eleves:
        voeu = voeux_map.get(eleve.id)
        finale = finales_map.get(eleve.id)

        choix_codes = []
        if voeu:
            choix_codes = [
                item.programme.code
                for item in voeu.choix.all().order_by("programme__ordre", "programme__code")
            ]

        rows.append({
            "matricule": eleve.identifiant or "-",
            "nom": eleve.nom or "-",
            "classe": eleve.classe.nom if eleve.classe else "-",
            "choix_eleve": ", ".join(choix_codes) if choix_codes else "-",
            "decision_administration": finale.programme_final.code if finale and finale.programme_final else "-",
        })

    return orientation, rows


def export_rapport_decision_finale_excel(ecole, annee_scolaire, niveau, classe, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport final"

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    ws.merge_cells("A1:E1")
    ws["A1"] = "Rapport de décision finale"
    ws["A1"].font = white_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A3"] = "Année scolaire"
    ws["B3"] = annee_scolaire.nom
    ws["C3"] = "Niveau"
    ws["D3"] = niveau.nom
    ws["A4"] = "Classe"
    ws["B4"] = classe.nom

    for cell in ["A3", "C3", "A4"]:
        ws[cell].font = bold_font

    headers = [
        "Matricule",
        "Nom de l'élève",
        "Classe",
        "Choix de l'élève",
        "Décision de l'administration",
    ]

    header_row = 6
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_row = header_row + 1
    for item in rows:
        ws.cell(row=data_row, column=1, value=item["matricule"])
        ws.cell(row=data_row, column=2, value=item["nom"])
        ws.cell(row=data_row, column=3, value=item["classe"])
        ws.cell(row=data_row, column=4, value=item["choix_eleve"])
        ws.cell(row=data_row, column=5, value=item["decision_administration"])
        data_row += 1

    widths = {
        "A": 22,
        "B": 30,
        "C": 20,
        "D": 28,
        "E": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"rapport_decision_finale_{annee_scolaire.nom}_{niveau.nom}_{classe.nom}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

@login_required
def rapport_decision_finale_page(request):
    ecole = get_ecole_user(request)
    form = RapportDecisionFinaleFilterForm(request.GET or None, ecole=ecole)

    rows = []
    orientation = None
    annee_scolaire = None
    niveau = None
    classe = None
    total_choix = 0
    total_decisions = 0

    if form.is_valid():
        annee_scolaire = form.cleaned_data["annee_scolaire"]
        niveau = form.cleaned_data["niveau"]
        classe = form.cleaned_data["classe"]

        orientation, rows = build_rapport_decision_finale_rows(
            ecole=ecole,
            annee_scolaire=annee_scolaire,
            niveau=niveau,
            classe=classe
        )

        total_choix = sum(1 for row in rows if row["choix_eleve"] != "-")
        total_decisions = sum(1 for row in rows if row["decision_administration"] != "-")

        if request.GET.get("export") == "excel" and orientation:
            return export_rapport_decision_finale_excel(
                ecole=ecole,
                annee_scolaire=annee_scolaire,
                niveau=niveau,
                classe=classe,
                rows=rows
            )

    context = {
        "form": form,
        "rows": rows,
        "orientation": orientation,
        "annee_scolaire": annee_scolaire,
        "niveau": niveau,
        "classe": classe,
        "total_choix": total_choix,
        "total_decisions": total_decisions,
    }
    return render(request, "orientation/rapport_decision_finale_page.html", context)