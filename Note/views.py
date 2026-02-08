from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from Ecole_admin.models import Note, Eleve, Matier, Classe
from Ecole_admin.utils.mixins import ActiveYearMixin, EcoleAssignMixin , UserAssignMixin
from Ecole_admin.utils.utils import get_annee_active


from django.http import HttpResponse
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime



from django.http import FileResponse
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont





from django.shortcuts import render, get_object_or_404


# -----------------------------------------------------------------
# ETAPE 1: ISOLER LA LOGIQUE DE CALCUL (HELPER)
# (C'est la logique de notre réponse précédente, mais dans une fonction)
# -----------------------------------------------------------------
def _calculer_bulletin_data(eleve, trimestre_num):
    """
    Calcule les données du bulletin pour un élève et un trimestre donnés.
    Retourne un dictionnaire (contexte) prêt à être utilisé par le template.
    """
    # 1. Récupérer les infos de base
    classe = eleve.classe
    annee_scolaire = eleve.annee_scolaire
    ecole = eleve.ecole
    
    # 2. Vérifier si l'élève a une classe/année
    if not classe or not annee_scolaire:
        return {'error': f"L'élève {eleve.nom} n'est pas assigné à une classe ou une année."}

    # 3. Récupérer les matières et les notes
    matieres_de_la_classe = Matier.objects.filter(classe=classe)
    notes_eleve = Note.objects.filter(
        eleve=eleve,
        annee_scolaire=annee_scolaire,
        trimestre=trimestre_num
    )

    # 4. Calculer les moyennes
    lignes_bulletin = []
    total_points_general = 0
    total_coeff_general = 0

    for matiere in matieres_de_la_classe:
        notes_pour_cette_matiere = notes_eleve.filter(matiere=matiere)
        
        points_matiere = sum(n.note * n.coefficient for n in notes_pour_cette_matiere)
        coeff_matiere = sum(n.coefficient for n in notes_pour_cette_matiere)
        
        moyenne_matiere = 0
        if coeff_matiere > 0:
            moyenne_matiere = points_matiere / coeff_matiere
        
        lignes_bulletin.append({
            'matiere': matiere.nom,
            'notes': notes_pour_cette_matiere,
            'total_points_matiere': round(points_matiere, 2),
            'total_coeff_matiere': coeff_matiere,
            'moyenne_matiere': round(moyenne_matiere, 2),
        })
        
        total_points_general += points_matiere
        total_coeff_general += coeff_matiere

    moyenne_generale = 0
    if total_coeff_general > 0:
        moyenne_generale = total_points_general / total_coeff_general

    trimestre_nom = dict(Note.TRIMESTRE_CHOISE).get(trimestre_num)

    # 5. Retourner le dictionnaire de contexte
    return {
        'eleve': eleve,
        'classe': classe,
        'ecole': ecole,
        'annee_scolaire': annee_scolaire,
        'trimestre_nom': trimestre_nom,
        'lignes_bulletin': lignes_bulletin,
        'total_points_general': round(total_points_general, 2),
        'total_coeff_general': total_coeff_general,
        'moyenne_generale': round(moyenne_generale, 2),
    }


def selectionner_bulletin(request):
    classes = Classe.objects.filter(ecole=request.user.ecole)
    eleves = Eleve.objects.filter(ecole=request.user.ecole,annee_scolaire=get_annee_active(request))
    context = {
        'classes': classes,
        'eleves': eleves,
        'trimestres': Note.TRIMESTRE_CHOISE,
        'bulletin_data_list': [],
        'error': None,
        'selected_classe_id': None,
        'selected_eleve_id': None,
        'selected_trimestre': None,
    }

    if request.method == 'POST':
        classe_id = request.POST.get('classe_id')
        eleve_id = request.POST.get('eleve_id')
        trimestre_str = request.POST.get('trimestre')

        context['selected_classe_id'] = int(classe_id) if classe_id and classe_id.isdigit() else None
        context['selected_eleve_id'] = int(eleve_id) if eleve_id else None
        context['selected_trimestre'] = int(trimestre_str) if trimestre_str else None

        if not trimestre_str:
            context['error'] = "Veuillez sélectionner un trimestre."
            return render(request, 'bulletin/selection.html', context)
        
        if not classe_id and not eleve_id:
            context['error'] = "Veuillez sélectionner une classe OU un élève."
            return render(request, 'bulletin/selection.html', context)

        trimestre_num = int(trimestre_str)
        eleves_a_traiter = []

        if eleve_id:
            try:
                eleve = Eleve.objects.get(id=eleve_id)
                eleves_a_traiter.append(eleve)
            except Eleve.DoesNotExist:
                context['error'] = "L'élève sélectionné n'existe pas."
        elif classe_id:
            try:
                classe = Classe.objects.get(id=classe_id)
                eleves_a_traiter = classe.eleves.all().order_by('nom')
                if not eleves_a_traiter:
                    context['error'] = "Cette classe ne contient aucun élève."
            except Classe.DoesNotExist:
                context['error'] = "La classe sélectionnée n'existe pas."

        bulletin_list = []
        for eleve in eleves_a_traiter:
            bulletin_data = _calculer_bulletin_data(eleve, trimestre_num)
            bulletin_list.append(bulletin_data)

        if len(bulletin_list) == 1:
            buffer = _generer_bulletin_pdf(bulletin_list[0])
            filename = f"bulletin_{bulletin_list[0]['eleve'].nom}.pdf"
            return FileResponse(buffer, as_attachment=True, filename=filename)
        
        context['bulletin_data_list'] = bulletin_list

        # 🔽 Vérifie si l'utilisateur a cliqué sur le bouton PDF
        if 'pdf' in request.POST and eleve_id:
            return _generer_bulletin_pdf(eleve, trimestre_num)

    return render(request, 'choisir_bulletin.html', context)


from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

def _generer_bulletin_pdf(bulletin_data):
    """
    Génère un PDF du bulletin à partir des données calculées.
    Retourne un objet BytesIO utilisable dans FileResponse.
    """
    buffer = BytesIO()

    # --- Enregistrement d'une police claire et lisible ---
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))  # police claire

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    # Styles personnalisés
    style_normal = ParagraphStyle(
        'normal',
        parent=styles['Normal'],
        fontName='HeiseiKakuGo-W5',
        fontSize=10.5,
        leading=15,             # interligne
        spaceAfter=6
    )
    style_center = ParagraphStyle(
        'center',
        parent=style_normal,
        alignment=TA_CENTER,
        fontSize=12,
        spaceAfter=10,
        letterSpacing=0.5       # espacement entre lettres
    )

    elements = []

    # --- En-tête avec logo gauche ---
    logo_path = getattr(bulletin_data['ecole'], 'logo', None)
    if logo_path:
        logo = Image(logo_path.path, width=60, height=60)
    else:
        logo = Paragraph("<b>LOGO</b>", style_center)

    gauche = Paragraph(
        """<b>RÉPUBLIQUE DE DJIBOUTI<br/>
        MINISTÈRE DE L'ÉDUCATION NATIONALE<br/>
        ET DE LA FORMATION PROFESSIONNELLE</b>""",
        style_normal
    )

    droite = Paragraph(
        f"""<b>Année scolaire :</b> {bulletin_data['annee_scolaire']}<br/>
        <b>École :</b> {bulletin_data['ecole'].nom}<br/>
        <b>Classe :</b> {bulletin_data['classe'].nom}""",
        style_normal
    )

    table_header = Table([[gauche, logo, droite]], colWidths=[220, 80, 220])
    table_header.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(table_header)
    elements.append(Spacer(1, 12))

    # --- Titre ---
    elements.append(Paragraph("<b><u>BULLETIN SCOLAIRE</u></b>", style_center))
    elements.append(Spacer(1, 10))

    # --- Informations élève ---
    eleve = bulletin_data['eleve']
    infos = [
        ["Nom :", f"{eleve.nom}", "Classe :", f"{bulletin_data['classe'].nom}"],
        ["Trimestre :", f"{bulletin_data['trimestre_nom']}", "Année :", f"{bulletin_data['annee_scolaire']}"],
    ]
    table_infos = Table(infos, colWidths=[90, 180, 90, 150])
    table_infos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'HeiseiKakuGo-W5'),
        ('FONTSIZE', (0, 0), (-1, -1), 10.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table_infos)
    elements.append(Spacer(1, 15))

    # --- Tableau des matières ---
    data = [["Matière", "Notes / 20", "Coef", "Total"]]
    for ligne in bulletin_data['lignes_bulletin']:
        notes_str = ", ".join(str(n.note) for n in ligne['notes'])
        data.append([
            ligne['matiere'],
            notes_str,
            ligne['total_coeff_matiere'],
            ligne['total_points_matiere']
        ])
    data.append(["", "", "Moyenne générale", f"{bulletin_data['moyenne_generale']}"])

    table_matieres = Table(data, colWidths=[200, 90, 60, 80])
    table_matieres.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1), 'HeiseiKakuGo-W5'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table_matieres)
    elements.append(Spacer(1, 15))

    # --- Appréciation ---
    elements.append(Paragraph(
        f"<b>Appréciation :</b> {bulletin_data.get('appreciation', 'Bon travail, continuez vos efforts !')}",
        style_normal
    ))
    elements.append(Spacer(1, 25))

    # --- Signatures ---
    # Récupération depuis la base de données (si disponible)

    signature_dir = getattr(bulletin_data['ecole'], 'signature', None)


    dir_cell = Image(signature_dir.path, width=80, height=40) if signature_dir else Paragraph("(cachet et signature)", style_center)

    table_sign = Table([
        ["Le Professeur Principal", "", "Le Directeur"],
        [ dir_cell]
    ], colWidths=[200, 100, 200])
    table_sign.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table_sign)

    # --- Génération finale ---
    doc.build(elements)
    buffer.seek(0)
    return buffer


class NoteListeView( LoginRequiredMixin ,ListView):
    model = Note
    template_name = 'note_liste.html'
    context_object_name = 'notes'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        
        if user.is_proffesseur:
            queryset = Note.objects.filter(
                user = user,
                ecole=self.request.user.ecole,
                annee_scolaire=get_annee_active(self.request)
            )
        else:
            queryset = Note.objects.filter(
                ecole=self.request.user.ecole,
                annee_scolaire=get_annee_active(self.request)
            )
        eleve = self.request.GET.get("eleve")
        trimestre = self.request.GET.get("trimestre")
        classe = self.request.GET.get("classe")


        if eleve and eleve != "":
            queryset = queryset.filter(eleve__id=eleve)
        if trimestre and trimestre != "":
            queryset = queryset.filter(trimestre__icontains=trimestre)
        if classe and classe != "":
            queryset = queryset.filter(eleve__classe_id=classe)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context["eleves"] = Eleve.objects.filter(classe__professeurs=self.request.user.proffeseur , ecole=self.request.user.ecole,annee_scolaire=get_annee_active(self.request)).all
        context["classes"] = Classe.objects.filter(professeurs=self.request.user.proffeseur ,ecole=self.request.user.ecole)
        return context


class NoteCreateView(UserAssignMixin , ActiveYearMixin , EcoleAssignMixin  , CreateView):
    model = Note
    template_name = 'note_create.html'
    fields = [
        'eleve',
        'matiere',
        'trimestre',
        'note',

    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        annee_active = get_annee_active(self.request)
        ecole = getattr(self.request.user, "ecole", None)
        prof = getattr(self.request.user, "proffeseur")

        if ecole and annee_active:
            form.fields["eleve"].queryset = Eleve.objects.filter(
                ecole=ecole,
                classe__professeurs = prof,
                annee_scolaire=annee_active
            )
            form.fields["matiere"].queryset = Matier.objects.filter(
                ecole=ecole,
                classe__professeurs = prof,
            )
        else:
            form.fields["eleve"].queryset = Eleve.objects.none()
            form.fields["matiere"].queryset = Matier.objects.none()

        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['eleves'] = Eleve.objects.filter(ecole=self.request.user.ecole,annee_scolaire=get_annee_active(self.request))
        context['matiers'] = Matier.objects.filter(ecole=self.request.user.ecole)
        context['Title'] = 'Ajouter une note'
        context['submit_text'] = 'Ajouter'
        return context

    def get_initial(self):
        initial = super().get_initial()
        eleve_id = self.kwargs.get("eleve_id")
        if eleve_id:
            initial["eleve"] = eleve_id
        return initial



    def get_success_url(self):
        eleve_id = self.kwargs.get("eleve_id")
        if eleve_id:
            return reverse("detaille", kwargs={"id": eleve_id})
        return reverse("ListeDesNotes")

class NoteUpdateView(UpdateView):
    model = Note
    template_name = 'note_create.html'
    success_url = reverse_lazy('ListeDesNotes')
    fields = [
        'eleve',
        'matiere',
        'periode',
        'note',

    ]


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['eleves'] = Eleve.objects.all()
        context['matiers'] = Matier.objects.all()
        context['Title'] = 'Modifier une note'
        context['submit_text'] = 'Modifier'
        return context


class NoteDeleteView(DeleteView):
    model = Note
    template_name = 'note_delete.html'
    success_url = reverse_lazy('ListeDesNotes')




import csv
from decimal import Decimal
from io import TextIOWrapper

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from Ecole_admin.models import (
    PeriodeScolaire, Devoir, Note, DispenseMatiere,
    Eleve, Classe, Matier
)
from Ecole_admin.form import (
    PeriodeScolaireForm, DevoirForm, NotesFilterForm,
    NoteSaisieSetupForm, DispenseMatiereForm, BulletinForm
)
from .services import compute_bulletin






from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404

from Ecole_admin.models import PeriodeScolaire
from Ecole_admin.form import PeriodeScolaireForm
from Ecole_admin.utils.periode import sync_periodes_auto, is_periode_usable

def _ctx_school(request):
    ecole = request.user.ecole
    from Ecole_admin.utils.utils import get_annee_active
    annee = get_annee_active(request)
    return ecole, annee


def _is_prof(user):
    return getattr(user, "role", None) == "proffesseur"


def _is_admin_or_secretaire(user):
    # adapte selon ton projet si tu as déjà une fonction
    return user.is_superuser or getattr(user, "role", "") in ("admin", "secretaire")

@login_required
def periode_list(request):
    ecole, annee = _ctx_school(request)

    # auto désactivation si fin dépassée / clôturée
    sync_periodes_auto(ecole, annee)

    periodes = PeriodeScolaire.objects.filter(ecole=ecole, annee_scolaire=annee).order_by("debut")
    form = PeriodeScolaireForm()

    if request.method == "POST":
        if not _is_admin_or_secretaire(request.user):
            return HttpResponse("Accès refusé", status=403)

        form = PeriodeScolaireForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.annee_scolaire = annee

            # si on coche est_active lors de création => rendre toutes les autres inactives
            if obj.est_active:
                PeriodeScolaire.objects.filter(ecole=ecole, annee_scolaire=annee).update(est_active=False)

            obj.save()
            messages.success(request, "Période enregistrée.")
            return redirect("periode_list")

    return render(request, "periode_list.html", {
        "periodes": periodes,
        "form": form,
    })

@login_required
def periode_toggle_active(request, pk):
    """
    ✅ active manuellement une période (mais interdit si clôturée/terminée)
    """
    if not _is_admin_or_secretaire(request.user):
        return HttpResponse("Accès refusé", status=403)

    ecole, annee = _ctx_school(request)
    p = get_object_or_404(PeriodeScolaire, pk=pk, ecole=ecole, annee_scolaire=annee)

    # interdit d'activer si inutilisable
    if not is_periode_usable(ecole, annee, p):
        messages.error(request, "Impossible : période clôturée ou terminée.")
        return redirect("periode_list")

    PeriodeScolaire.objects.filter(ecole=ecole, annee_scolaire=annee).update(est_active=False)
    p.est_active = True
    p.save(update_fields=["est_active"])
    messages.success(request, "Période active mise à jour.")
    return redirect("periode_list")


# ==========================
# DEVOIRS
# ==========================
@login_required
def devoir_list(request):
    ecole, annee = _ctx_school(request)

    qs = Devoir.objects.filter(ecole=ecole, annee_scolaire=annee).select_related("niveau", "professeur", "matiere", "periode")
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(matiere__nom__icontains=q) | Q(professeur__nom_conplet__icontains=q))

    return render(request, "devoir_list.html", {"devoirs": qs.order_by("-date_creation"), "q": q})


@login_required
def devoir_create(request):
    if not (_is_admin_or_secretaire(request.user) or _is_prof(request.user)):
        return HttpResponse("Accès refusé", status=403)

    ecole, annee = _ctx_school(request)

    if request.method == "POST":
        form = DevoirForm(request.POST, ecole=ecole, annee_scolaire=annee)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.ecole = ecole
            obj.annee_scolaire = annee
            obj.save()
            form.save_m2m()
            messages.success(request, "Devoir créé.")
            return redirect("devoir_list")
    else:
        form = DevoirForm(ecole=ecole, annee_scolaire=annee)

    return render(request, "devoir_create.html", {"form": form})


@login_required
def devoir_detail(request, pk):
    ecole, annee = _ctx_school(request)
    d = get_object_or_404(
        Devoir.objects.select_related("niveau", "professeur", "matiere", "periode").prefetch_related("classes"),
        pk=pk, ecole=ecole, annee_scolaire=annee
    )
    return render(request, "devoir_detail.html", {"d": d})


# ==========================
# GESTION DES NOTES (LIST)
# ==========================

def periode_to_trimestre(periode) -> int:
    nom = (getattr(periode, "nom", "") or "").lower()
    if "3" in nom:
        return 3
    if "2" in nom:
        return 2
    return 1




@login_required
def notes_gestion(request):
    ecole, annee = _ctx_school(request)

    form = NotesFilterForm(request.GET or None, ecole=ecole, annee_scolaire=annee)

    qs = (
        Note.objects
        .filter(ecole=ecole, annee_scolaire=annee)
        .select_related("eleve", "eleve__classe", "matiere", "devoir", "devoir__periode")
        .order_by("-id")
    )

    if form.is_valid():
        classe = form.cleaned_data.get("classe")
        matiere = form.cleaned_data.get("matiere")
        devoir = form.cleaned_data.get("devoir")
        periode = form.cleaned_data.get("periode")
        q = form.cleaned_data.get("q")

        if classe:
            qs = qs.filter(eleve__classe=classe)
        if matiere:
            qs = qs.filter(matiere=matiere)
        if devoir:
            qs = qs.filter(devoir=devoir)
        if periode:
            qs = qs.filter(devoir__periode=periode)
        if q:
            qs = qs.filter(Q(eleve__nom__icontains=q) | Q(eleve__identifiant__icontains=q))

    return render(request, "notes_gestion.html", {
        "form": form,
        "notes": qs[:800],
    })



@login_required
def note_saisie_setup(request):
    if not (_is_admin_or_secretaire(request.user) or _is_prof(request.user)):
        return HttpResponse("Accès refusé", status=403)

    ecole, annee = _ctx_school(request)
    form = NoteSaisieSetupForm(request.GET or None, ecole=ecole, annee_scolaire=annee)

    if request.method == "GET" and form.is_valid():
        devoir = form.cleaned_data["devoir"]
        classe = form.cleaned_data["classe"]

        if not devoir.classes.filter(pk=classe.pk).exists():
            messages.error(request, "Cette classe n'est pas associée à ce devoir.")
            return redirect("note_saisie_setup")

        return redirect("note_saisie", devoir_id=devoir.id, classe_id=classe.id)

    return render(request, "note_saisie_setup.html", {"form": form})



@login_required
def note_saisie(request, devoir_id, classe_id):
    if not (_is_admin_or_secretaire(request.user) or _is_prof(request.user)):
        return HttpResponse("Accès refusé", status=403)

    ecole, annee = _ctx_school(request)
    devoir = get_object_or_404(Devoir, pk=devoir_id, ecole=ecole, annee_scolaire=annee)
    classe = get_object_or_404(Classe, pk=classe_id, ecole=ecole)

    if not devoir.classes.filter(pk=classe.pk).exists():
        return HttpResponseBadRequest("Classe non autorisée pour ce devoir.")

    eleves = Eleve.objects.filter(ecole=ecole, annee_scolaire=annee, classe=classe).order_by("nom")

    exist = {
        n.eleve_id: n
        for n in Note.objects.filter(ecole=ecole, annee_scolaire=annee, devoir=devoir, eleve__in=eleves)
    }

    if request.method == "POST":
        tri = periode_to_trimestre(devoir.periode)

        with transaction.atomic():
            for e in eleves:
                key_note = f"note_{e.id}"
                key_coef = f"coef_{e.id}"

                val = (request.POST.get(key_note) or "").strip()
                coef = (request.POST.get(key_coef) or "").strip()

                if val == "":
                    continue

                try:
                    note_val = Decimal(val.replace(",", "."))
                except Exception:
                    continue

                try:
                    coef_val = int(coef) if coef else devoir.coefficient
                except Exception:
                    coef_val = devoir.coefficient

                obj = exist.get(e.id)
                if obj:
                    obj.note = note_val
                    obj.coefficient = coef_val
                    obj.matiere = devoir.matiere
                    obj.user = request.user
                    obj.trimestre = tri
                    obj.save()
                else:
                    Note.objects.create(
                        user=request.user,
                        eleve=e,
                        matiere=devoir.matiere,
                        devoir=devoir,
                        trimestre=tri,
                        note=note_val,
                        coefficient=coef_val,
                        annee_scolaire=annee,
                        ecole=ecole,
                    )

        messages.success(request, "Notes enregistrées.")
        return redirect("notes_gestion")

    return render(request, "note_saisie.html", {
        "devoir": devoir,
        "classe": classe,
        "eleves": eleves,
        "exist": exist,
    })



@login_required
def note_import(request):
    if not (_is_admin_or_secretaire(request.user) or _is_prof(request.user)):
        return HttpResponse("Accès refusé", status=403)

    import csv
    from io import TextIOWrapper
    from decimal import Decimal
    import openpyxl
    from django.db import transaction

    ecole, annee = _ctx_school(request)

    def periode_to_trimestre(periode) -> int:
        nom = (getattr(periode, "nom", "") or "").lower()
        if "3" in nom:
            return 3
        if "2" in nom:
            return 2
        return 1

    devoirs = Devoir.objects.filter(ecole=ecole, annee_scolaire=annee).order_by("-date_creation")

    # résultats affichés dans template
    ctx = {"devoirs": devoirs, "result": None}

    if request.method != "POST":
        return render(request, "note_import.html", ctx)

    devoir_id = request.POST.get("devoir_id")
    fichier = request.FILES.get("fichier")

    if not devoir_id or not fichier:
        ctx["result"] = {
            "ok": 0, "ko": 0, "rows": 0,
            "errors": ["Sélectionner un devoir ET un fichier (CSV ou XLSX). Vérifie aussi enctype=multipart/form-data."]
        }
        return render(request, "note_import.html", ctx)

    devoir = get_object_or_404(Devoir, pk=devoir_id, ecole=ecole, annee_scolaire=annee)

    # si devoir.periode est vide, on met trimestre=1
    tri = periode_to_trimestre(getattr(devoir, "periode", None))

    filename = (fichier.name or "").lower()
    ok, ko, rows = 0, 0, 0
    errors = []

    def norm_ident(x: str) -> str:
        x = ("" if x is None else str(x)).strip()
        if x.endswith(".0"):
            x = x[:-2]
        return x

    def handle_row(ident, note_s, coef_s):
        nonlocal ok, ko, rows, errors
        rows += 1

        ident = norm_ident(ident)
        note_s = ("" if note_s is None else str(note_s)).strip()
        coef_s = ("" if coef_s is None else str(coef_s)).strip()

        if not ident or not note_s:
            ko += 1
            if len(errors) < 15:
                errors.append(f"Ligne {rows}: identifiant ou note vide.")
            return

        # ✅ NE PAS filtrer sur annee_scolaire ici (tu peux avoir des élèves d’une autre année en DB)
        eleve = Eleve.objects.filter(ecole=ecole, identifiant=ident).first()
        if not eleve:
            ko += 1
            if len(errors) < 15:
                errors.append(f"Ligne {rows}: élève introuvable (identifiant={ident}).")
            return

        try:
            note_val = Decimal(note_s.replace(",", "."))
        except Exception:
            ko += 1
            if len(errors) < 15:
                errors.append(f"Ligne {rows}: note invalide ({note_s}).")
            return

        try:
            coef_val = int(coef_s) if coef_s else devoir.coefficient
        except Exception:
            coef_val = devoir.coefficient

        Note.objects.update_or_create(
            ecole=ecole,
            annee_scolaire=annee,
            devoir=devoir,
            eleve=eleve,
            defaults={
                "user": request.user,
                "matiere": devoir.matiere,
                "note": note_val,
                "coefficient": coef_val,
                "trimestre": tri,
            }
        )
        ok += 1

    try:
        with transaction.atomic():
            # XLSX
            if filename.endswith(".xlsx"):
                fichier.seek(0)
                wb = openpyxl.load_workbook(fichier, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))

                if not all_rows:
                    ctx["result"] = {"ok": 0, "ko": 0, "rows": 0, "errors": ["Fichier Excel vide."]}
                    return render(request, "note_import.html", ctx)

                # détecter header ou pas
                start_index = 0
                h = [str(x).strip().lower() for x in (all_rows[0] or [])]
                if any(k in h for k in ["identifiant", "note", "coefficient"]):
                    start_index = 1

                for r in all_rows[start_index:]:
                    ident = r[0] if len(r) > 0 else ""
                    note_s = r[1] if len(r) > 1 else ""
                    coef_s = r[2] if len(r) > 2 else ""
                    handle_row(ident, note_s, coef_s)

            # CSV
            else:
                wrapper = TextIOWrapper(fichier.file, encoding="utf-8")
                reader = csv.DictReader(wrapper)

                # si colonnes pas bonnes → aucune ligne utile
                if not reader.fieldnames:
                    ctx["result"] = {"ok": 0, "ko": 0, "rows": 0, "errors": ["CSV invalide ou vide."]}
                    return render(request, "note_import.html", ctx)

                # on accepte aussi majuscules/minuscules
                for row in reader:
                    ident = row.get("identifiant") or row.get("Identifiant") or row.get("IDENTIFIANT")
                    note_s = row.get("note") or row.get("Note") or row.get("NOTE")
                    coef_s = row.get("coefficient") or row.get("Coefficient") or row.get("COEFFICIENT")
                    handle_row(ident, note_s, coef_s)

    except Exception as ex:
        ctx["result"] = {"ok": ok, "ko": ko, "rows": rows, "errors": [f"Erreur import: {ex}"] + errors}
        return render(request, "note_import.html", ctx)

    ctx["result"] = {"ok": ok, "ko": ko, "rows": rows, "errors": errors}
    return render(request, "note_import.html", ctx)



@login_required
def note_import_template_xlsx(request):
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Import_Notes"

    # Header EXACT attendu
    ws.append(["identifiant", "note", "coefficient"])

    # Exemple
    ws.append(["2026010700", 15, 1])
    ws.append(["2026010701", 13, 2])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modele_import_notes.xlsx"'
    wb.save(response)
    return response



# ==========================
# DISPENSES
# ==========================
@login_required
def dispense_list(request):
    if not (_is_admin_or_secretaire(request.user) or _is_prof(request.user)):
        return HttpResponse("Accès refusé", status=403)

    ecole, annee = _ctx_school(request)

    qs = DispenseMatiere.objects.filter(ecole=ecole, annee_scolaire=annee).select_related("eleve", "matiere", "periode").order_by("-created_at")

    form = DispenseMatiereForm(request.POST or None, ecole=ecole, annee_scolaire=annee)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.ecole = ecole
        obj.annee_scolaire = annee
        obj.save()
        messages.success(request, "Dispense enregistrée.")
        return redirect("dispense_list")

    return render(request, "dispense_matiere.html", {"items": qs[:500], "form": form})


# ==========================
# BULLETIN
# ==========================
@login_required
def bulletin_list(request):
    ecole, annee = _ctx_school(request)
    form = BulletinForm(request.GET or None, ecole=ecole, annee_scolaire=annee)

    result = None
    eleve = None
    periode = None

    if form.is_valid() and request.GET:
        eleve = form.cleaned_data["eleve"]
        periode = form.cleaned_data["periode"]
        result = compute_bulletin(eleve, periode, annee, ecole)

    return render(request, "bulletin_list.html", {
        "form": form,
        "result": result,
        "eleve": eleve,
        "periode": periode,
    })


from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from datetime import timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Sum, F, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import render, get_object_or_404
from django.utils import timezone

from Ecole_admin.models import (
    AnneeScolaire, Niveau, Classe, Eleve, Matier, Proffeseur,
    Note, Absence, PeriodeScolaire,
    AppreciationPeriode, AppreciationAbsence, DecisionAbsence
)

# -------------------------
# Helpers
# -------------------------
def _q2(x) -> Decimal:
    """Decimal arrondi 2 décimales."""
    if x is None:
        return Decimal("0.00")
    if not isinstance(x, Decimal):
        x = Decimal(str(x))
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def _periode_to_trimestre(periode: PeriodeScolaire) -> int:
    """
    Ton modèle Note utilise trimestre = 1/2/3 (pas FK vers PeriodeScolaire).
    On déduit le numéro à partir du nom 'Trimestre 1/2/3' sinon fallback via ordre par date.
    """
    # cas: "Trimestre 1"
    for n in (1, 2, 3):
        if str(n) in (periode.nom or ""):
            return n
    # fallback: basé sur l'ordre chronologique des périodes de l'année
    qs = PeriodeScolaire.objects.filter(ecole=periode.ecole, annee_scolaire=periode.annee_scolaire).order_by("debut")
    ids = list(qs.values_list("id", flat=True))
    if periode.id in ids:
        idx = ids.index(periode.id)  # 0..2
        return min(3, max(1, idx + 1))
    return 1

def appreciation_from_moyenne(ecole, annee, moyenne: Decimal) -> str:
    row = (AppreciationPeriode.objects
           .filter(ecole=ecole, annee_scolaire=annee, actif=True,
                   note_min__lte=moyenne, note_max__gte=moyenne)
           .order_by("note_min")
           .first())
    return row.nom if row else ""

def appreciation_absence_from_count(ecole, annee, abs_count: int) -> str:
    row = (AppreciationAbsence.objects
           .filter(ecole=ecole, annee_scolaire=annee, actif=True,
                   abs_min__lte=abs_count, abs_max__gte=abs_count)
           .order_by("abs_min")
           .first())
    return row.nom if row else ""

def decision_absence_from_count(ecole, annee, abs_count: int) -> str:
    row = (DecisionAbsence.objects
           .filter(ecole=ecole, annee_scolaire=annee, actif=True,
                   max_abs__lte=abs_count)
           .order_by("-max_abs")
           .first())
    return row.get_statut_display() if row else ""

def _teacher_name_for_matiere(classe: Classe, matiere: Matier) -> str:
    """
    Chez toi: Proffeseur.matieres (FK Matier) + Proffeseur.classes (M2M).
    """
    prof = (Proffeseur.objects
            .filter(matieres=matiere, classes=classe, actif=True)
            .order_by("nom_conplet")
            .first())
    return prof.nom_conplet if prof else "-"

def _moyenne_eleve_matiere(eleve: Eleve, matiere: Matier, trimestre: int, annee: AnneeScolaire, ecole) -> tuple[Decimal, int]:
    """
    moyenne pondérée par coefficient des notes: sum(note*coef)/sum(coef)
    retourne (moyenne, nb_notes)
    """
    qs = Note.objects.filter(
        eleve=eleve, matiere=matiere, trimestre=trimestre,
        annee_scolaire=annee, ecole=ecole
    )

    nb = qs.count()
    if nb == 0:
        return Decimal("0.00"), 0

    agg = qs.aggregate(
        s_num=Coalesce(Sum(F("note") * F("coefficient"), output_field=DecimalField()), Decimal("0")),
        s_den=Coalesce(Sum("coefficient"), 0),
    )
    den = agg["s_den"] or 0
    if den == 0:
        return Decimal("0.00"), nb
    return _q2(Decimal(agg["s_num"]) / Decimal(den)), nb

def _moyenne_classe_matiere(classe: Classe, matiere: Matier, trimestre: int, annee: AnneeScolaire, ecole) -> Decimal:
    """
    moyenne de la classe (moyenne des moyennes élèves pour cette matière)
    (approche fiable et simple)
    """
    eleves = Eleve.objects.filter(classe=classe, annee_scolaire=annee, ecole=ecole)
    moys = []
    for e in eleves:
        m, nb = _moyenne_eleve_matiere(e, matiere, trimestre, annee, ecole)
        if nb > 0:
            moys.append(m)
    if not moys:
        return Decimal("0.00")
    return _q2(sum(moys, Decimal("0")) / Decimal(len(moys)))

def _rang_eleve_matiere(classe: Classe, eleve: Eleve, matiere: Matier, trimestre: int, annee: AnneeScolaire, ecole) -> str:
    """
    Rang de l'élève dans la classe pour la matière (basé sur moyenne matière).
    """
    eleves = list(Eleve.objects.filter(classe=classe, annee_scolaire=annee, ecole=ecole))
    scores = []
    for e in eleves:
        m, nb = _moyenne_eleve_matiere(e, matiere, trimestre, annee, ecole)
        # si pas de note => ignore du classement
        if nb > 0:
            scores.append((e.id, m))
    if not scores:
        return "-"
    scores.sort(key=lambda t: t[1], reverse=True)  # desc
    # rang dense (1,2,2,4) ? ici rang simple (1..n) basé sur position
    ids_sorted = [x[0] for x in scores]
    if eleve.id not in ids_sorted:
        return "-"
    return str(ids_sorted.index(eleve.id) + 1)

def _moyenne_generale_eleve(eleve: Eleve, trimestre: int, annee: AnneeScolaire, ecole) -> Decimal:
    """
    Moyenne générale: pondérée sur toutes les notes (toutes matières) par coefficient.
    """
    qs = Note.objects.filter(eleve=eleve, trimestre=trimestre, annee_scolaire=annee, ecole=ecole)
    agg = qs.aggregate(
        s_num=Coalesce(Sum(F("note") * F("coefficient"), output_field=DecimalField()), Decimal("0")),
        s_den=Coalesce(Sum("coefficient"), 0),
    )
    den = agg["s_den"] or 0
    if den == 0:
        return Decimal("0.00")
    return _q2(Decimal(agg["s_num"]) / Decimal(den))

def _absences_heures_periode(eleve: Eleve, periode: PeriodeScolaire, annee: AnneeScolaire, ecole) -> tuple[Decimal, int]:
    """
    Template affiche: absences_heures.
    On calcule sur Absence.statut='absence' dans [debut..fin].
    Si h_debut/h_fin => heures réelles, sinon 1h par enregistrement.
    Retourne (heures_decimal, nb_absences)
    """
    qs = Absence.objects.filter(
        eleve=eleve,
        annee_scolaire=annee,
        ecole=ecole,
        statut="absence",
        date__range=(periode.debut, periode.fin),
    )

    nb = qs.count()
    total_hours = Decimal("0")

    for a in qs:
        if a.h_debut and a.h_fin:
            dt1 = datetime.combine(a.date, a.h_debut)
            dt2 = datetime.combine(a.date, a.h_fin)
            diff = dt2 - dt1
            hours = Decimal(str(diff.total_seconds() / 3600))
            if hours < 0:
                hours = Decimal("0")
            total_hours += hours
        else:
            total_hours += Decimal("1")  # fallback 1h

    return _q2(total_hours), nb


@dataclass
class BulletinRow:
    eleve: Eleve
    data: dict


@login_required
def bulletins_visual(request):
    """
    Rend le template fourni sans changer le visuel.
    Construit `bulletins` conforme à bulletins_visual.html.
    """
    user = request.user
    ecole = user.ecole
    today = timezone.localdate()

    # filtres
    annees = AnneeScolaire.objects.all().order_by("-debut")
    annee_id = request.GET.get("annee")
    annee = AnneeScolaire.objects.filter(id=annee_id).first() if annee_id else AnneeScolaire.get_active()

    niveaux = Niveau.objects.filter(ecole=ecole, actif=True).order_by("ordre", "nom")
    niveau_id = request.GET.get("niveau") or ""
    classes = Classe.objects.filter(ecole=ecole, actif=True)
    if niveau_id:
        classes = classes.filter(niveau_id=niveau_id)
    classes = classes.order_by("niveau__ordre", "nom")

    classe_id = request.GET.get("classe") or ""
    periodes = PeriodeScolaire.objects.filter(ecole=ecole, annee_scolaire=annee).order_by("debut")
    periode_id = request.GET.get("periode") or ""
    eleve_id = request.GET.get("eleve") or "all"

    bulletins = []
    periode_obj = None
    classe_obj = None

    if classe_id and periode_id:
        classe_obj = get_object_or_404(Classe, id=classe_id, ecole=ecole)
        periode_obj = get_object_or_404(PeriodeScolaire, id=periode_id, ecole=ecole, annee_scolaire=annee)
        trimestre = _periode_to_trimestre(periode_obj)

        # élèves
        eleves_qs = Eleve.objects.filter(classe=classe_obj, annee_scolaire=annee, ecole=ecole).order_by("nom")
        if eleve_id != "all":
            eleves_qs = eleves_qs.filter(id=eleve_id)

        # matières liées à la classe
        matieres = Matier.objects.filter(classe=classe_obj, ecole=ecole).order_by("nom")

        for eleve in eleves_qs:
            lignes = []

            # lignes par matière
            for m in matieres:
                moy_eleve, nb_notes = _moyenne_eleve_matiere(eleve, m, trimestre, annee, ecole)
                moy_classe = _moyenne_classe_matiere(classe_obj, m, trimestre, annee, ecole)
                rang = _rang_eleve_matiere(classe_obj, eleve, m, trimestre, annee, ecole)
                enseignant = _teacher_name_for_matiere(classe_obj, m)
                appreciation = appreciation_from_moyenne(ecole, annee, moy_eleve) if nb_notes > 0 else ""

                lignes.append({
                    "matiere": m.nom,
                    "enseignant": enseignant,
                    "nb_note": nb_notes,
                    "rang": rang,
                    "moy_eleve": f"{moy_eleve}" if nb_notes > 0 else "-",
                    "moy_classe": f"{moy_classe}" if nb_notes > 0 else "-",
                    "appreciation": appreciation,
                })

            # moyenne générale période
            moy_gen = _moyenne_generale_eleve(eleve, trimestre, annee, ecole)
            app_global = appreciation_from_moyenne(ecole, annee, moy_gen) if moy_gen > 0 else ""

            # absences
            abs_heures, abs_count = _absences_heures_periode(eleve, periode_obj, annee, ecole)
            app_abs = appreciation_absence_from_count(ecole, annee, abs_count)
            dec_abs = decision_absence_from_count(ecole, annee, abs_count)

            # on respecte ton template: b.data.absences_heures et b.data.appreciation_globale
            # tu veux aussi que absences s'appliquent: on combine en texte clair sans changer ton visuel
            # (tu peux ajuster le texte comme tu veux)
            appreciation_globale = app_global
            if app_abs:
                appreciation_globale = f"{app_global} | Abs: {app_abs}" if app_global else f"Abs: {app_abs}"

            bulletins.append(BulletinRow(
                eleve=eleve,
                data={
                    "lignes": lignes,
                    "moyenne_generale": f"{moy_gen}" if moy_gen > 0 else "-",
                    "absences_heures": f"{abs_heures}",
                    "appreciation_globale": appreciation_globale,
                    # optionnel si plus tard tu veux afficher quelque part:
                    "decision_absence": dec_abs,
                    "nb_absences": abs_count,
                }
            ))

    # dropdown élèves dépend de classe
    eleves_dropdown = Eleve.objects.filter(classe_id=classe_id, annee_scolaire=annee, ecole=ecole).order_by("nom") if classe_id else Eleve.objects.none()

    return render(request, "bulletins_visual.html", {
        "ecole": ecole,
        "today": today,

        "annees": annees,
        "annee": annee,
        "annee_id": str(annee.id) if annee else "",

        "niveaux": niveaux,
        "niveau_id": str(niveau_id),

        "classes": classes,
        "classe_id": str(classe_id),

        "periodes": periodes,
        "periode": periode_obj,
        "periode_id": str(periode_id),

        "eleves": eleves_dropdown,
        "eleve_id": str(eleve_id),

        "bulletins": bulletins,
    })




from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.http import HttpResponse
from Ecole_admin.models import AnneeScolaire
from Ecole_admin.models import Niveau, Classe, PeriodeScolaire
from .services_rapport import build_rapport_classe

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


@login_required
def rapport_conseil_classe(request):
    ecole = request.user.ecole

    # ✅ Toutes les années (pas de ecole dans ton modèle)
    annees = AnneeScolaire.objects.all().order_by("-debut")

    # ✅ année choisie sinon active
    annee_id = request.GET.get("annee") or ""
    if annee_id.isdigit():
        annee = AnneeScolaire.objects.filter(pk=int(annee_id)).first()
    else:
        annee = AnneeScolaire.get_active()

    if not annee:
        annee = annees.first()

    # Filtres
    niveaux = Niveau.objects.filter(ecole=ecole, actif=True).order_by("ordre", "nom")
    periodes = PeriodeScolaire.objects.filter(ecole=ecole, annee_scolaire=annee).order_by("debut")
    classes_all = Classe.objects.filter(ecole=ecole, actif=True).select_related("niveau").order_by("niveau__ordre", "nom")

    niveau_id = request.GET.get("niveau") or "all"
    classe_id = request.GET.get("classe") or ""
    periode_id = request.GET.get("periode") or ""

    classes = classes_all
    if niveau_id != "all" and niveau_id.isdigit():
        classes = classes_all.filter(niveau_id=int(niveau_id))

    classes_footer = classes_all if niveau_id == "all" else classes

    selected_classe = classes_all.filter(pk=classe_id).first() if classe_id.isdigit() else None
    selected_periode = PeriodeScolaire.objects.filter(pk=periode_id, ecole=ecole, annee_scolaire=annee).first() if periode_id.isdigit() else None

    rapport = None
    if selected_classe and selected_periode:
        rapport = build_rapport_classe(selected_classe, selected_periode, annee, ecole)

    # ✅ EXPORT EXCEL
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    if request.GET.get("export") == "excel":
        if not (rapport and selected_classe and selected_periode):
            return HttpResponse("Choisir classe et période avant l'export Excel.", content_type="text/plain")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"

        # Titre
        title = f"Rapport général d'examen - {selected_classe.nom} - {selected_periode.nom} - {annee.nom}"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        # largeur merge (colonnes = 6 fixes + matières + 6 final)
        total_cols = 6 + len(rapport["subjects"]) + 6
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

        # Entêtes
        headers = ["N°", "Nom", "Classe", "Date naissance", "Sexe", "Statut"]
        for m in rapport["subjects"]:
            headers.append(m.nom.upper())
        headers += ["Total", "Nb matières", "Moyenne", "Absences (h)", "Appréciation", "Rang"]

        row_start = 3
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_start, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Données
        for i, r in enumerate(rapport["rows"], start=1):
            rr = row_start + i
            ws.cell(rr, 1, i)
            ws.cell(rr, 2, r["eleve"].nom)
            ws.cell(rr, 3, r["eleve"].classe.nom)
            ws.cell(rr, 4, str(r["eleve"].date_naissancce))
            ws.cell(rr, 5, r.get("sexe", "-"))
            ws.cell(rr, 6, r.get("status", "-"))

            c = 7
            # ✅ IMPORTANT: notes_list (pas r["notes"])
            for idx_m in range(len(rapport["subjects"])):
                val = r["notes_list"][idx_m]
                ws.cell(rr, c, float(val) if val is not None else "")
                c += 1

            ws.cell(rr, c, float(r["total"])); c += 1
            ws.cell(rr, c, int(r["nb"])); c += 1
            ws.cell(rr, c, float(r["moyenne"])); c += 1
            ws.cell(rr, c, float(r.get("absences", 0))); c += 1
            ws.cell(rr, c, r.get("appreciation", "")); c += 1
            ws.cell(rr, c, int(r["rang"]))

        # Largeurs
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        filename = f"rapport_{selected_classe.nom}_{selected_periode.nom}_{annee.nom}.xlsx".replace(" ", "_")
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


    ctx = {
        "ecole": ecole,
        "annees": annees,
        "annee": annee,
        "annee_id": str(annee.id) if annee else "",

        "niveaux": niveaux,
        "classes": classes,
        "classes_footer": classes_footer,
        "periodes": periodes,

        "niveau_id": str(niveau_id),
        "classe_id": str(classe_id),
        "periode_id": str(periode_id),

        "selected_classe": selected_classe,
        "selected_periode": selected_periode,
        "rapport": rapport,
        "today": timezone.localdate(),
    }
    return render(request, "rapport_conseil_classe.html", ctx)
